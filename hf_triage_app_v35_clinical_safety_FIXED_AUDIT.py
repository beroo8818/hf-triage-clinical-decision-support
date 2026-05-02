import csv
import json
import os
import requests
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import ttkbootstrap as tb
except Exception:
    class _FallbackWindow(tk.Tk):
        def __init__(self, themename=None, *args, **kwargs):
            super().__init__(*args, **kwargs)

    class _FallbackBootstrap:
        Window = _FallbackWindow

    tb = _FallbackBootstrap()

try:
    from openpyxl import Workbook, load_workbook
    HAS_OPENPYXL = True
except Exception:
    Workbook = None
    load_workbook = None
    HAS_OPENPYXL = False

APP_TITLE = "HF Triage Clinical Safety & Decision Support Assistant"
N8N_EMERGENCY_WEBHOOK_URL = "https://beroo90.app.n8n.cloud/webhook/hf-triage-emergency-alert"
N8N_DOCTOR_DECISION_LOOKUP_URL = "https://beroo90.app.n8n.cloud/webhook/hf-get-doctor-decision"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "hf_triage_data_v3.json")


def _resolve_runtime_data_file(filename):
    """Prefer a writable Documents folder, then fall back to the app folder."""
    candidates = [
        os.path.join(os.path.expanduser("~"), "Documents", "HF_Triage_Project", filename),
        os.path.join(BASE_DIR, filename),
    ]
    for candidate in candidates:
        try:
            os.makedirs(os.path.dirname(candidate), exist_ok=True)
            with open(candidate, "a", encoding="utf-8"):
                pass
            return candidate
        except Exception:
            continue
    return os.path.join(BASE_DIR, filename)


AUDIT_FILE = _resolve_runtime_data_file("hf_triage_audit_log_v1.csv")
ENTRY_POINT = "Emergency Department"
WORKFLOW_IMAGE_FILE = os.path.join(BASE_DIR, "workflow_presentation.png")

UI_COLORS = {
    "navy": "#0f172a",
    "blue": "#1d4ed8",
    "blue_soft": "#eaf2ff",
    "teal_soft": "#ecfeff",
    "green_soft": "#ecfdf3",
    "amber_soft": "#fffbeb",
    "red_soft": "#fef2f2",
    "panel": "#fbfdff",
    "border": "#d9e3f0",
    "text": "#0f172a",
    "muted": "#475569",
}

AUDIT_FIELDNAMES = [
    "event_id",
    "timestamp",
    "event_type",
    "user_role",
    "patient_id",
    "patient_name",
    "target_output",
    "risk_output",
    "risk_score",
    "final_pathway",
    "recommended_action",
    "clinician_action",
    "override_risk",
    "override_pathway",
    "override_reason",
    "document_type",
    "document_status",
    "confidence_level",
    "trigger_summary",
    "outcome_status",
    "notes",
    "local_ed_acuity",
    "safety_lock_status",
    "clinician_confirmation_status",
    "emergency_gates",
]

DOCUMENT_TYPES = ["SOAP Note", "Discharge Summary", "Referral Note"]
CLINICIAN_ROLES = ["Triage Nurse", "Emergency Physician", "Cardiology Registrar", "Ward Physician"]

ALERT_SUPPRESSION_WINDOW_MINUTES = 30
ALERT_DELIVERY_RULES = {
    "High Risk": "Interruptive alert",
    "Medium Risk": "Non-interruptive dashboard flag",
    "Low Risk": "Silent monitoring only",
    "Blocked - Critical Data Missing": "Safety block - no final triage",
    "Blocked - Invalid Vitals": "Safety block - invalid data",
}

# -----------------------------------------------------------------------------
# Clinical safety layer
# -----------------------------------------------------------------------------
# These ranges are input-sanity boundaries, not diagnostic cutoffs. They catch
# impossible or unsafe entries before the CDSS output is trusted.
VITAL_HARD_RANGES = {
    "age": (0, 120, "Age"),
    "systolic_bp": (40, 300, "Systolic BP"),
    "diastolic_bp": (20, 180, "Diastolic BP"),
    "heart_rate": (20, 250, "Heart rate"),
    "respiratory_rate": (4, 80, "Respiratory rate"),
    "oxygen_saturation": (50, 100, "Oxygen saturation"),
    "temperature": (30.0, 45.0, "Temperature"),
}

CRITICAL_DATA_FIELDS = [
    "Age",
    "Systolic BP",
    "Diastolic BP",
    "Heart rate",
    "Respiratory rate",
    "Oxygen saturation",
    "Temperature",
]

LOCAL_ED_ACUITY_LEVELS = [
    "Level 1 - Resuscitation / Immediate",
    "Level 2 - Emergent",
    "Level 3 - Urgent",
    "Level 4 - Less urgent",
    "Level 5 - Non-urgent",
    "Blocked - Critical data required",
]

def send_emergency_alert_to_n8n(payload):
    """
    Send emergency/high-risk triage event to n8n.
    Keep patient data anonymized.
    """
    if not N8N_EMERGENCY_WEBHOOK_URL:
        return False

    try:
        response = requests.post(
            N8N_EMERGENCY_WEBHOOK_URL,
            json=payload,
            timeout=8
        )

        if response.status_code in (200, 201):
            print("Emergency alert sent to n8n successfully.")
            return True

        print(f"n8n returned status code: {response.status_code}")
        print(response.text)
        return False

    except Exception as error:
        print(f"n8n emergency alert failed: {error}")
        return False
def _yes(value):
    return str(value or "").strip().lower() == "yes"


def _coerce_number(value, as_float=False):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        number = float(text)
        return number if as_float else int(number)
    except (TypeError, ValueError):
        return None


def validate_vital_ranges_snapshot(snapshot):
    """Pure validation helper used by the app and unit tests."""
    errors = []
    for key, (minimum, maximum, label) in VITAL_HARD_RANGES.items():
        raw_value = snapshot.get(key, "")
        if raw_value in (None, ""):
            continue
        value = _coerce_number(raw_value, as_float=(key == "temperature"))
        if value is None:
            errors.append(f"{label} must be numeric")
            continue
        if value < minimum or value > maximum:
            errors.append(f"{label} value {raw_value} is outside allowed range {minimum}-{maximum}")
    return errors


def evaluate_emergency_gates_snapshot(snapshot):
    """Pure emergency gate helper used by the app and unit tests."""
    gates = []
    sbp = _coerce_number(snapshot.get("systolic_bp"))
    hr = _coerce_number(snapshot.get("heart_rate"))
    rr = _coerce_number(snapshot.get("respiratory_rate"))
    spo2 = _coerce_number(snapshot.get("oxygen_saturation"))

    if _yes(snapshot.get("severe_dyspnea")):
        gates.append("Severe dyspnea / marked respiratory distress")
    if _yes(snapshot.get("confusion")):
        gates.append("Confusion / altered mental status")
    if _yes(snapshot.get("pulmonary_edema_signs")):
        gates.append("Suspected acute pulmonary oedema")
    if _yes(snapshot.get("arrhythmia_symptoms")):
        gates.append("Possible unstable arrhythmia symptoms")
    if _yes(snapshot.get("chest_pain")):
        gates.append("Chest pain requiring immediate ED review")
    if sbp is not None and sbp < 90:
        gates.append(f"Hypotension: SBP {sbp} mmHg")
    if spo2 is not None and spo2 < 90:
        gates.append(f"Severe hypoxia: SpO2 {spo2}%")
    if rr is not None and rr >= 30:
        gates.append(f"Severe tachypnoea: RR {rr}/min")
    if hr is not None and hr >= 130:
        gates.append(f"Severe tachycardia: HR {hr} bpm")
    return gates


def determine_local_ed_acuity(emergency_gates, risk_output, missing_critical=None, invalid_ranges=None):
    missing_critical = missing_critical or []
    invalid_ranges = invalid_ranges or []
    if invalid_ranges:
        return "Blocked - Critical data required"
    if emergency_gates:
        shock_like_terms = ("Hypotension", "Severe hypoxia", "acute pulmonary", "altered mental", "Severe tachypnoea")
        if any(any(term in gate for term in shock_like_terms) for gate in emergency_gates):
            return "Level 1 - Resuscitation / Immediate"
        return "Level 2 - Emergent"
    if missing_critical:
        return "Blocked - Critical data required"
    if risk_output == "High Risk":
        return "Level 2 - Emergent"
    if risk_output == "Medium Risk":
        return "Level 3 - Urgent"
    if risk_output == "Low Risk":
        return "Level 4 - Less urgent"
    return "Blocked - Critical data required"


def discharge_is_locked(evidence, clinician_confirmed=False):
    """Pure discharge safety helper used by the app and unit tests."""
    reasons = []
    if evidence.get("emergency_gate_active"):
        reasons.append("emergency gate active")
    if evidence.get("missing_critical"):
        reasons.append("critical data missing")
    if evidence.get("invalid_ranges"):
        reasons.append("invalid vital-sign range")
    if evidence.get("risk_output") == "High Risk":
        reasons.append("high-risk output")
    if not clinician_confirmed:
        reasons.append("clinician confirmation pending")
    return (len(reasons) > 0, reasons)

PREDICTION_TARGETS = [
    "Risk of same-day clinical deterioration",
    "Risk of urgent escalation during the same visit",
    "Risk of 30-day hospitalization/readmission",
]

RISK_CLASSES = ["Low Risk", "Medium Risk", "High Risk", "Blocked - Critical Data Missing", "Blocked - Invalid Vitals"]
FINAL_PATHWAYS = [
    "Routine Evaluation",
    "Urgent Care",
    "Immediate ED Escalation",
    "ED High-Acuity Admission",
    "ICU/ICCU Consideration",
    "Ward Admission",
    "Admission",
    "Referral",
    "Discharge After Clinician Confirmation",
    "Final Triage Blocked",
]

PREDICTION_TIMING_TEXT = "After triage data entry and before final triage decision"
PREDICTION_USERS_TEXT = "Triage nurse and emergency physician"
PREDICTION_OUTPUT_TEXT = "Risk score + risk category classification + early warning alerts + recommended clinical actions"
CLINICIAN_OVERSIGHT_TEXT = "Final decision remains under clinician oversight"

PRESENTATION_CURRENT_WORKFLOW_ORDER = [
    "Start",
    "Patient Arrival",
    "Patient Registration / Waiting",
    "Initial Assessment",
    "Clinical Evaluation",
    "Medication Review",
    "Functional Assessment",
    "Triage Decision",
    "Admission to Ward",
    "ICU Admission",
    "Discharge",
    "Inconsistent Follow-up",
    "End",
]

PRESENTATION_CURRENT_WORKFLOW_SUGGESTIONS = {
    "Start": {
        "actor": "Process start",
        "input": "Patient presents for evaluation",
        "action": "Start the current clinical triage process",
        "output": "Workflow initiated",
    },
    "Patient Arrival": {
        "actor": "Emergency Department / Outpatient Clinic",
        "input": "Patient arrival at service entry point",
        "action": "Receive the patient in the emergency department or outpatient clinic",
        "output": "Patient available for registration and waiting",
    },
    "Patient Registration / Waiting": {
        "actor": "Registration desk / front office",
        "input": "Demographics, visit details, queue status",
        "action": "Register the patient and place them in the waiting flow",
        "output": "Registered patient awaiting assessment",
    },
    "Initial Assessment": {
        "actor": "Nurse",
        "input": "Initial complaint, visible condition, first vital signs",
        "action": "Perform the initial nursing assessment and collect first-line triage findings",
        "output": "Initial assessment profile",
    },
    "Clinical Evaluation": {
        "actor": "Physician",
        "input": "Initial assessment findings and bedside clinical exam",
        "action": "Carry out the physician clinical evaluation",
        "output": "Physician evaluation summary",
    },
    "Medication Review": {
        "actor": "Pharmacist",
        "input": "Current medications, prescriptions, adherence history",
        "action": "Review medication profile, treatment adherence, and medication-related concerns",
        "output": "Medication review summary",
    },
    "Functional Assessment": {
        "actor": "Physiotherapy",
        "input": "Mobility, exercise tolerance, symptom impact on activity",
        "action": "Assess functional status and daily activity limitation",
        "output": "Functional assessment summary",
    },
    "Triage Decision": {
        "actor": "Clinical team",
        "input": "Nursing, physician, pharmacist, and physiotherapy findings",
        "action": "Make the current manual triage decision",
        "output": "Admit, ICU, or discharge direction",
    },
    "Admission to Ward": {
        "actor": "Ward team",
        "input": "Admit branch from triage decision",
        "action": "Admit the patient to the ward",
        "output": "Ward admission",
    },
    "ICU Admission": {
        "actor": "ICU team",
        "input": "ICU branch from triage decision",
        "action": "Transfer and admit the patient to ICU",
        "output": "ICU admission",
    },
    "Discharge": {
        "actor": "Clinical team",
        "input": "Discharge branch from triage decision",
        "action": "Discharge the patient from the current encounter",
        "output": "Discharged patient",
    },
    "Inconsistent Follow-up": {
        "actor": "Outpatient follow-up pathway",
        "input": "Discharged patient with delayed or inconsistent follow-up plan",
        "action": "Continue follow-up outside the hospital pathway with variable consistency",
        "output": "Inconsistent follow-up outcome",
    },
    "End": {
        "actor": "Process end",
        "input": "Ward admission, ICU admission, or follow-up outcome",
        "action": "Close the current clinical triage workflow",
        "output": "End",
    },
}

PRESENTATION_AI_WORKFLOW_ORDER = [
    "Start",
    "Patient Arrival",
    "Digital Registration & Data",
    "Aggregate All Inputs",
    "AI Risk Stratification",
    "Real-Time AI Outputs",
    "Clinical Dashboard",
    "Vitals Unstable?",
    "Detect Red Flags",
    "High-risk Alert",
    "Immediate Physician Review",
    "Physician Confirms",
    "AI Risk High (>75)?",
    "AI Risk Medium (25-75)?",
    "Physician Review & Confirm",
    "ICU Admission",
    "IV Diuretics",
    "Urgent Cardio Review",
    "Ward Admission",
    "Monitoring Protocol",
    "Medication Optimization",
    "Pharmacist Validate Med",
    "Safe Discharge",
    "Early Follow-up",
    "Home Monitoring Plan",
    "Functional Decline?",
    "Physio Validate Rehab",
    "Cardiac Rehab Referral",
    "Track Outcomes & Update",
    "Clinician Override Reason",
    "End Process",
    "Retrain Model",
]

PRESENTATION_AI_WORKFLOW_SUGGESTIONS = {
    "Start": {
        "actor": "Process start",
        "input": "Patient arrives for AI-supported triage",
        "action": "Start the AI-clinical triage assistant workflow",
        "output": "AI triage process initiated",
    },
    "Patient Arrival": {
        "actor": "ED intake",
        "input": "Patient arrives",
        "action": "Receive the patient at triage entry",
        "output": "Patient ready for digital registration",
    },
    "Digital Registration & Data": {
        "actor": "Registration / digital intake",
        "input": "Demographics, registration, initial data capture",
        "action": "Digitally register the patient and capture core data",
        "output": "Structured registration data",
    },
    "Aggregate All Inputs": {
        "actor": "System / data integration layer",
        "input": "Collected triage, history, lab, and medication inputs",
        "action": "Aggregate all structured and available clinical inputs",
        "output": "Unified AI-ready input set",
    },
    "AI Risk Stratification": {
        "actor": "AI model",
        "input": "Unified AI-ready input set",
        "action": "Run AI risk stratification",
        "output": "AI risk scores and branch probabilities",
    },
    "Real-Time AI Outputs": {
        "actor": "AI model",
        "input": "AI risk stratification results",
        "action": "Publish real-time AI outputs",
        "output": "Dashboard-ready AI outputs",
    },
    "Clinical Dashboard": {
        "actor": "Clinical dashboard",
        "input": "Real-time AI outputs",
        "action": "Display risk results to the care team",
        "output": "Dashboard review state",
    },
    "Vitals Unstable?": {
        "actor": "AI logic / rules",
        "input": "Dashboard review state and vital-sign checks",
        "action": "Check whether the patient has unstable vitals",
        "output": "Yes: red flag trigger / No: continue AI risk path",
    },
    "Detect Red Flags": {
        "actor": "AI logic / rules",
        "input": "Yes: Red Flag Trigger",
        "action": "Detect critical red-flag findings",
        "output": "Red flags detected",
    },
    "High-risk Alert": {
        "actor": "AI alert engine",
        "input": "Red flags detected",
        "action": "Generate high-risk alert",
        "output": "Urgent high-risk notification",
    },
    "Immediate Physician Review": {
        "actor": "Physician",
        "input": "Urgent high-risk notification",
        "action": "Perform immediate physician review",
        "output": "Immediate review result",
    },
    "Physician Confirms": {
        "actor": "Physician",
        "input": "Immediate review result",
        "action": "Confirm the high-risk pathway",
        "output": "Confirmed high-risk decision",
    },
    "AI Risk High (>75)?": {
        "actor": "AI scoring logic",
        "input": "Dashboard review state when vitals are not unstable",
        "action": "Check if AI high-risk score is greater than 75",
        "output": "Yes / No",
    },
    "AI Risk Medium (25-75)?": {
        "actor": "AI scoring logic",
        "input": "No from AI high-risk branch",
        "action": "Check if AI medium-risk score is between 25 and 75",
        "output": "Yes / No",
    },
    "Physician Review & Confirm": {
        "actor": "Physician",
        "input": "AI risk branch requiring clinician confirmation",
        "action": "Review the AI recommendation and confirm the appropriate branch",
        "output": "Approved high / medium / low risk path",
    },
    "ICU Admission": {
        "actor": "ICU team",
        "input": "Approve high risk / physician confirms",
        "action": "Admit patient to ICU",
        "output": "ICU admission",
    },
    "IV Diuretics": {
        "actor": "Clinical team",
        "input": "ICU/high-risk management plan",
        "action": "Start IV diuretics",
        "output": "IV diuretic therapy started",
    },
    "Urgent Cardio Review": {
        "actor": "Cardiology",
        "input": "Post-ICU or high-risk branch",
        "action": "Conduct urgent cardiology review",
        "output": "Urgent cardiology assessment",
    },
    "Ward Admission": {
        "actor": "Ward team",
        "input": "Approved medium-risk branch",
        "action": "Admit patient to ward",
        "output": "Ward admission",
    },
    "Monitoring Protocol": {
        "actor": "Ward team",
        "input": "Ward admission",
        "action": "Apply monitoring protocol",
        "output": "Monitored inpatient pathway",
    },
    "Medication Optimization": {
        "actor": "Clinical team",
        "input": "Monitoring protocol and medication profile",
        "action": "Optimize medications",
        "output": "Medication optimization plan",
    },
    "Pharmacist Validate Med": {
        "actor": "Pharmacist",
        "input": "Medication optimization and medication alert check",
        "action": "Validate medication plan",
        "output": "Pharmacist-validated medication plan",
    },
    "Safe Discharge": {
        "actor": "Clinical team",
        "input": "Approved low-risk branch",
        "action": "Proceed with safe discharge",
        "output": "Discharged patient",
    },
    "Early Follow-up": {
        "actor": "Follow-up coordination",
        "input": "Discharged patient",
        "action": "Arrange early follow-up",
        "output": "Early follow-up scheduled",
    },
    "Home Monitoring Plan": {
        "actor": "Follow-up coordination",
        "input": "Early follow-up pathway",
        "action": "Create home monitoring plan",
        "output": "Home monitoring plan",
    },
    "Functional Decline?": {
        "actor": "AI / rehab review",
        "input": "Evaluate functional data",
        "action": "Check for functional decline",
        "output": "Yes / No functional decline",
    },
    "Physio Validate Rehab": {
        "actor": "Physiotherapy",
        "input": "Yes functional decline",
        "action": "Validate rehab need",
        "output": "Physio rehab validation",
    },
    "Cardiac Rehab Referral": {
        "actor": "Rehab / cardiology coordination",
        "input": "Approved rehab pathway",
        "action": "Create cardiac rehabilitation referral",
        "output": "Cardiac rehab referral",
    },
    "Track Outcomes & Update": {
        "actor": "Monitoring / analytics team",
        "input": "All downstream clinical outcomes",
        "action": "Track outcomes and update the system",
        "output": "Updated outcome data",
    },
    "Clinician Override Reason": {
        "actor": "Clinician",
        "input": "Override decision",
        "action": "Document the clinician override reason",
        "output": "Stored override rationale",
    },
    "End Process": {
        "actor": "Process end",
        "input": "Tracked outcomes and process completion",
        "action": "Close the AI-assisted triage process",
        "output": "End Process",
    },
    "Retrain Model": {
        "actor": "AI operations / data science",
        "input": "Tracked outcomes and updated functional data",
        "action": "Retrain the model from feedback loops",
        "output": "Updated model version",
    },
}

WORKFLOW_SUGGESTIONS = {
    **PRESENTATION_CURRENT_WORKFLOW_SUGGESTIONS,
    **PRESENTATION_AI_WORKFLOW_SUGGESTIONS,
}

DATA_ELEMENT_SUGGESTIONS = {
    "Shortness of breath": {
        "category": "Symptoms",
        "data_type": "Semi-structured",
        "source": "Triage form / nurse interview",
        "why_it_matters": "Core heart failure symptom and early severity indicator",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Edema": {
        "category": "Symptoms",
        "data_type": "Semi-structured",
        "source": "Triage form / nurse interview",
        "why_it_matters": "Suggests fluid overload",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Fatigue": {
        "category": "Symptoms",
        "data_type": "Semi-structured",
        "source": "Triage form / nurse interview",
        "why_it_matters": "Common symptom that supports the clinical picture",
        "priority": "Medium",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Chest pain": {
        "category": "Symptoms",
        "data_type": "Semi-structured",
        "source": "Triage form / nurse interview",
        "why_it_matters": "May indicate an urgent clinical condition",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Blood pressure": {
        "category": "Vital signs",
        "data_type": "Structured",
        "source": "Triage vital signs",
        "why_it_matters": "Detects hypotension or hemodynamic instability",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Heart rate": {
        "category": "Vital signs",
        "data_type": "Structured",
        "source": "Triage vital signs",
        "why_it_matters": "May reflect stress or arrhythmia",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Respiratory rate": {
        "category": "Vital signs",
        "data_type": "Structured",
        "source": "Triage vital signs",
        "why_it_matters": "Important for respiratory distress assessment",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Oxygen saturation": {
        "category": "Vital signs",
        "data_type": "Structured",
        "source": "Triage vital signs",
        "why_it_matters": "Critical severity marker",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Temperature": {
        "category": "Vital signs",
        "data_type": "Structured",
        "source": "Triage vital signs",
        "why_it_matters": "May indicate infection or trigger",
        "priority": "Medium",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Previous admissions": {
        "category": "Medical history",
        "data_type": "Structured",
        "source": "EMR",
        "why_it_matters": "Prior admission history increases risk concern",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Medication history": {
        "category": "Medical history",
        "data_type": "Semi-structured",
        "source": "EMR / patient interview",
        "why_it_matters": "Supports risk context and adherence review",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Comorbidities": {
        "category": "Medical history",
        "data_type": "Structured",
        "source": "EMR / problem list",
        "why_it_matters": "Adds important clinical context",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Severe dyspnea": {
        "category": "Red flags",
        "data_type": "Semi-structured",
        "source": "Triage assessment",
        "why_it_matters": "Major indicator for urgent review",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Hypotension": {
        "category": "Red flags",
        "data_type": "Structured",
        "source": "Blood pressure",
        "why_it_matters": "Suggests instability",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Low oxygen saturation": {
        "category": "Red flags",
        "data_type": "Structured",
        "source": "Pulse oximeter",
        "why_it_matters": "Indicates respiratory compromise",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Confusion": {
        "category": "Red flags",
        "data_type": "Semi-structured",
        "source": "Nurse observation",
        "why_it_matters": "May indicate severe physiological stress",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Pulmonary edema signs": {
        "category": "Red flags",
        "data_type": "Semi-structured",
        "source": "Clinical assessment",
        "why_it_matters": "Suggests acute decompensation",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Arrhythmia symptoms": {
        "category": "Red flags",
        "data_type": "Semi-structured",
        "source": "Patient interview / triage note",
        "why_it_matters": "May point to unstable cardiac status",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Triage nurse notes": {
        "category": "Clinical notes",
        "data_type": "Unstructured",
        "source": "Nurse note",
        "why_it_matters": "Captures context not in checkbox fields",
        "priority": "High",
        "available_at_triage": "Yes",
        "status": "Keep",
    },
    "Referral notes": {
        "category": "Clinical notes",
        "data_type": "Unstructured",
        "source": "Referral document",
        "why_it_matters": "Provides outside clinical context",
        "priority": "Medium",
        "available_at_triage": "Sometimes",
        "status": "Keep",
    },
    "Previous discharge summary": {
        "category": "Clinical notes",
        "data_type": "Unstructured",
        "source": "EMR",
        "why_it_matters": "Useful for recent course and previous deterioration",
        "priority": "Medium",
        "available_at_triage": "Sometimes",
        "status": "Keep",
    },
}

REQUIRED_WORKFLOW_STEPS = PRESENTATION_CURRENT_WORKFLOW_ORDER
REQUIRED_DATA_ELEMENTS = list(DATA_ELEMENT_SUGGESTIONS.keys())


def build_workflow_steps_from_order(step_order, suggestion_map):
    steps = []
    for index, step_name in enumerate(step_order, start=1):
        suggestion = suggestion_map[step_name]
        steps.append({
            "step_number": index,
            "step_name": step_name,
            "actor": suggestion["actor"],
            "input": suggestion["input"],
            "action": suggestion["action"],
            "output": suggestion["output"],
        })
    return steps


def default_workflow_steps(workflow_type="current"):
    if workflow_type == "ai":
        return build_workflow_steps_from_order(PRESENTATION_AI_WORKFLOW_ORDER, PRESENTATION_AI_WORKFLOW_SUGGESTIONS)
    return build_workflow_steps_from_order(PRESENTATION_CURRENT_WORKFLOW_ORDER, PRESENTATION_CURRENT_WORKFLOW_SUGGESTIONS)


def default_data_inventory():
    items = []
    for name, suggestion in DATA_ELEMENT_SUGGESTIONS.items():
        items.append({
            "category": suggestion["category"],
            "data_element": name,
            "data_type": suggestion["data_type"],
            "source": suggestion["source"],
            "why_it_matters": suggestion["why_it_matters"],
            "priority": suggestion["priority"],
            "available_at_triage": suggestion["available_at_triage"],
            "status": suggestion["status"],
        })
    return items


def default_patient_cases():
    return []


class LocalStorage:
    def __init__(self, filename=DATA_FILE):
        self.filename = filename

    def delete_file(self):
        if os.path.exists(self.filename):
            os.remove(self.filename)

    def load_data(self):
        if not os.path.exists(self.filename):
            data = {
                "project_target": PREDICTION_TARGETS[1],
                "workflow_steps": default_workflow_steps(),
                "data_inventory": default_data_inventory(),
                "patient_cases": default_patient_cases(),
            }
            self.save_data(data)
            return data

        try:
            with open(self.filename, "r", encoding="utf-8") as file:
                data = json.load(file)
        except Exception:
            data = {
                "project_target": PREDICTION_TARGETS[1],
                "workflow_steps": default_workflow_steps(),
                "data_inventory": default_data_inventory(),
                "patient_cases": default_patient_cases(),
            }

        data.setdefault("project_target", PREDICTION_TARGETS[1])
        data.setdefault("workflow_steps", default_workflow_steps())
        data.setdefault("data_inventory", default_data_inventory())
        data.setdefault("patient_cases", default_patient_cases())
        return data

    def save_data(self, data):
        with open(self.filename, "w", encoding="utf-8") as file:
            json.dump(data, file, indent=4, ensure_ascii=False)


class HFTriageApp(tb.Window):
    def __init__(self):
        super().__init__(themename="litera")
        self.title(APP_TITLE)
        self.geometry("1680x920")
        self.minsize(1440, 780)
        self.option_add("*Font", "{Segoe UI} 9")

        self.storage = LocalStorage()
        self.app_data = self.storage.load_data()
        self.workflow_steps = self.app_data["workflow_steps"]
        self.data_inventory = self.app_data["data_inventory"]
        self.patient_cases = self.app_data["patient_cases"]
        self.ensure_audit_file()

        self.selected_workflow_index = None
        self.selected_data_index = None
        self.selected_case_index = None
        self.alert_state_by_patient = {}
        self.current_alert_context = {}
        self._patient_load_in_progress = False
        self._patient_id_to_index = {}

        self.configure_styles()
        self.create_widgets()
        self.refresh_workflow_tree()
        self.refresh_data_tree()
        self.refresh_case_tree()
        self.update_status("Ready")

    def configure_styles(self):
        style = ttk.Style(self)
        style.configure("Treeview", rowheight=26, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 9, "bold"))
        style.configure("Header.TLabel", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("SectionHint.TLabel", font=("Segoe UI", 8), foreground=UI_COLORS["muted"])
        style.configure("TLabelframe.Label", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("TNotebook.Tab", padding=(14, 6), font=("Segoe UI Semibold", 10, "bold"))
        style.configure("ConfidenceHigh.TLabel", foreground="#166534", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("ConfidenceModerate.TLabel", foreground="#b45309", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("ConfidenceLow.TLabel", foreground="#b91c1c", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("ConfidenceDefault.TLabel", foreground="#334155", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("RiskHigh.TLabel", foreground="#b91c1c", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("RiskMedium.TLabel", foreground="#b45309", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("RiskLow.TLabel", foreground="#166534", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("RiskDefault.TLabel", foreground="#334155", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("HumanLoop.TLabel", foreground="#0f766e", font=("Segoe UI Semibold", 10, "bold"))
        style.configure("PatientField.TLabel", font=("Segoe UI", 8))
        style.configure("PredictiveLabel.TLabel", font=("Segoe UI Semibold", 9, "bold"))
        style.configure("PredictiveValue.TLabel", font=("Segoe UI", 9))
        style.configure("PredictiveValueAccent.TLabel", font=("Segoe UI", 9), foreground="#0f766e")
        style.configure("Compact.TButton", font=("Segoe UI", 8), padding=(8, 4))
        style.configure("Compact.TCombobox", font=("Segoe UI", 8), padding=(2, 2))
        style.configure("Compact.TEntry", font=("Segoe UI", 8), padding=(2, 2))
        self.option_add("*TCombobox*Listbox.Font", "{Segoe UI} 8")

    def create_widgets(self):
        hero_frame = self.create_banner(
            self,
            "HF Triage Clinical Decision Support Workspace",
            "A cleaner capstone prototype for triage review, clinician override, structured documentation, and audit-ready traceability.",
            bg=UI_COLORS["navy"],
        )
        hero_frame.pack(fill="x", padx=8, pady=(8, 6))

        top_frame = ttk.Frame(self, padding=(8, 4, 8, 4))
        top_frame.pack(fill="x")

        ttk.Label(top_frame, text="Project Prediction Target:", style="Header.TLabel").pack(side="left", padx=(0, 8))
        self.project_target_var = tk.StringVar(value=self.app_data.get("project_target", PREDICTION_TARGETS[1]))
        self.project_target_combo = ttk.Combobox(top_frame, textvariable=self.project_target_var, values=PREDICTION_TARGETS, state="readonly", width=40, style="Compact.TCombobox")
        self.project_target_combo.pack(side="left")
        self.project_target_combo.bind("<<ComboboxSelected>>", lambda event: self.save_project_target())

        ttk.Button(top_frame, text="Save All", style="Compact.TButton", command=self.save_all_data).pack(side="right", padx=4)
        ttk.Button(top_frame, text="Reload Local Data", style="Compact.TButton", command=self.reload_local_data).pack(side="right", padx=4)
        ttk.Button(top_frame, text="Clear Saved Cases", style="Compact.TButton", command=self.clear_saved_patient_cases).pack(side="right", padx=4)

        quick_scroll_frame = ttk.Frame(self, padding=(8, 0, 8, 6))
        quick_scroll_frame.pack(fill="x")
        ttk.Label(quick_scroll_frame, text="Quick Scroll:", style="Header.TLabel").pack(side="left", padx=(0, 4))
        ttk.Button(quick_scroll_frame, text="Top", style="Compact.TButton", command=lambda: self.patient_canvas.yview_moveto(0)).pack(side="left", padx=1)
        ttk.Button(quick_scroll_frame, text="Use Case", style="Compact.TButton", command=lambda: self.scroll_patient_to_widget(self.patient_sections["Predictive Use Case"])).pack(side="left", padx=1)
        ttk.Button(quick_scroll_frame, text="Form", style="Compact.TButton", command=lambda: self.scroll_patient_to_widget(self.patient_sections["Patient Case Form"])).pack(side="left", padx=1)
        ttk.Button(quick_scroll_frame, text="Insights", style="Compact.TButton", command=lambda: self.scroll_patient_to_widget(self.patient_sections["Decision Support Insights"])).pack(side="left", padx=1)
        ttk.Button(quick_scroll_frame, text="Summary", style="Compact.TButton", command=lambda: self.scroll_patient_to_widget(self.patient_sections["Case Summary"])).pack(side="left", padx=1)
        ttk.Button(quick_scroll_frame, text="Explain", style="Compact.TButton", command=lambda: self.scroll_patient_to_widget(self.patient_sections["Explainability Panel"])).pack(side="left", padx=1)
        ttk.Button(quick_scroll_frame, text="Docs", style="Compact.TButton", command=lambda: self.scroll_patient_to_widget(self.patient_sections["Clinical Documentation"])).pack(side="left", padx=1)
        ttk.Button(quick_scroll_frame, text="Saved", style="Compact.TButton", command=lambda: self.scroll_patient_to_widget(self.patient_sections["Saved Patient Cases"])).pack(side="left", padx=1)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=6, pady=(0, 6))

        self.phase1_tab = ttk.Frame(self.notebook, padding=8)
        self.phase2_tab = ttk.Frame(self.notebook, padding=8)
        self.audit_tab = ttk.Frame(self.notebook, padding=8)

        self.patient_tab_container = ttk.Frame(self.notebook)
        self.patient_canvas = tk.Canvas(self.patient_tab_container, highlightthickness=0)
        self.patient_scrollbar = ttk.Scrollbar(self.patient_tab_container, orient="vertical", command=self.patient_canvas.yview)
        self.patient_tab = ttk.Frame(self.patient_canvas, padding=8)

        self.patient_canvas.configure(yscrollcommand=self.patient_scrollbar.set)
        self.patient_scrollbar.pack(side="right", fill="y")
        self.patient_canvas.pack(side="left", fill="both", expand=True)
        self.patient_canvas_window = self.patient_canvas.create_window((0, 0), window=self.patient_tab, anchor="nw")

        def _configure_patient_scroll_region(event=None):
            self.patient_canvas.configure(scrollregion=self.patient_canvas.bbox("all"))

        def _configure_patient_canvas_width(event):
            self.patient_canvas.itemconfigure(self.patient_canvas_window, width=event.width)

        self.patient_tab.bind("<Configure>", _configure_patient_scroll_region)
        self.patient_canvas.bind("<Configure>", _configure_patient_canvas_width)

        def _on_mousewheel(event):
            try:
                if event.delta:
                    self.patient_canvas.yview_scroll(int(-5 * (event.delta / 120)), "units")
                elif getattr(event, 'num', None) == 4:
                    self.patient_canvas.yview_scroll(-5, "units")
                elif getattr(event, 'num', None) == 5:
                    self.patient_canvas.yview_scroll(5, "units")
            except Exception:
                pass

        def _bind_mousewheel(event=None):
            self.patient_canvas.bind_all("<MouseWheel>", _on_mousewheel)
            self.patient_canvas.bind_all("<Button-4>", _on_mousewheel)
            self.patient_canvas.bind_all("<Button-5>", _on_mousewheel)

        def _unbind_mousewheel(event=None):
            self.patient_canvas.unbind_all("<MouseWheel>")
            self.patient_canvas.unbind_all("<Button-4>")
            self.patient_canvas.unbind_all("<Button-5>")

        self.patient_canvas.bind("<Enter>", _bind_mousewheel)
        self.patient_canvas.bind("<Leave>", _unbind_mousewheel)

        self.notebook.add(self.phase1_tab, text="Phase 1 – Workflow Builder")
        self.notebook.add(self.phase2_tab, text="Phase 2 – Data Inventory Builder")
        self.notebook.add(self.patient_tab_container, text="Patient Case Entry")
        self.notebook.add(self.audit_tab, text="Phase 4 – Audit Log")

        self.build_phase1_tab()
        self.build_phase2_tab()
        self.build_patient_tab()
        self.build_audit_tab()
        self.refresh_audit_tree()

        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w", padding=4)
        status_bar.pack(fill="x", side="bottom")
        self.bind_all("<Prior>", lambda e: self.patient_canvas.yview_scroll(-1, "page"))
        self.bind_all("<Next>", lambda e: self.patient_canvas.yview_scroll(1, "page"))
        self.bind_all("<Home>", lambda e: self.patient_canvas.yview_moveto(0))
        self.bind_all("<End>", lambda e: self.patient_canvas.yview_moveto(1))

    def scroll_patient_to_widget(self, widget):
        self.update_idletasks()
        total_height = max(1, self.patient_tab.winfo_height())
        y = max(0, widget.winfo_y() - 10)
        self.patient_canvas.yview_moveto(y / total_height)

    def _bind_scroll_to_children(self, parent):
        def _scroll_from_child(event):
            try:
                if getattr(event, "delta", 0):
                    self.patient_canvas.yview_scroll(int(-5 * (event.delta / 120)), "units")
                elif getattr(event, 'num', None) == 4:
                    self.patient_canvas.yview_scroll(-5, "units")
                elif getattr(event, 'num', None) == 5:
                    self.patient_canvas.yview_scroll(5, "units")
            except Exception:
                pass
            return "break"

        for child in parent.winfo_children():
            if isinstance(child, (tk.Text, ttk.Entry, ttk.Combobox, ttk.Treeview, tk.Listbox, ttk.Frame, ttk.LabelFrame, ttk.Label, ttk.Button, tk.Canvas)):
                child.bind("<MouseWheel>", _scroll_from_child, add="+")
                child.bind("<Button-4>", _scroll_from_child, add="+")
                child.bind("<Button-5>", _scroll_from_child, add="+")
            self._bind_scroll_to_children(child)

    def update_status(self, text):
        self.status_var.set(text)

    def create_banner(self, parent, title, subtitle, bg="#0f172a", fg="#ffffff"):
        frame = tk.Frame(parent, bg=bg, padx=18, pady=14, highlightthickness=1, highlightbackground=bg)
        title_label = tk.Label(frame, text=title, bg=bg, fg=fg, font=("Segoe UI Semibold", 16, "bold"), anchor="w")
        title_label.pack(anchor="w")
        subtitle_label = tk.Label(frame, text=subtitle, bg=bg, fg="#dbeafe", font=("Segoe UI", 9), anchor="w", justify="left")
        subtitle_label.pack(anchor="w", pady=(4, 0))
        return frame

    def create_metric_card(self, parent, title, value_var, subtitle="", bg="#ffffff", value_fg=None, width=220):
        frame = tk.Frame(parent, bg=bg, padx=14, pady=12, highlightthickness=1, highlightbackground=UI_COLORS["border"])
        frame.pack_propagate(False)
        frame.configure(width=width, height=88)
        tk.Label(frame, text=title, bg=bg, fg=UI_COLORS["muted"], font=("Segoe UI Semibold", 8, "bold"), anchor="w").pack(anchor="w")
        value_label = tk.Label(frame, textvariable=value_var, bg=bg, fg=value_fg or UI_COLORS["text"], font=("Segoe UI Semibold", 15, "bold"), anchor="w")
        value_label.pack(anchor="w", pady=(4, 2))
        if subtitle:
            tk.Label(frame, text=subtitle, bg=bg, fg=UI_COLORS["muted"], font=("Segoe UI", 8), anchor="w", justify="left").pack(anchor="w")
        return frame, value_label

    def apply_text_panel_style(self, widget):
        try:
            widget.configure(
                bg=UI_COLORS["panel"],
                fg=UI_COLORS["text"],
                relief="flat",
                insertbackground=UI_COLORS["text"],
                padx=10,
                pady=8,
                highlightthickness=1,
                highlightbackground=UI_COLORS["border"],
            )
        except Exception:
            pass

    def update_patient_dashboard_cards(self, evidence=None):
        if not hasattr(self, "patient_kpi_vars"):
            return
        if evidence is None:
            evidence = {"missing_critical": []}
        current_risk = self.output_risk_var.get().strip() if hasattr(self, "output_risk_var") else "Not calculated"
        current_pathway = self.output_pathway_var.get().strip() if hasattr(self, "output_pathway_var") else "Not calculated"
        current_confidence = self.confidence_var.get().strip() if hasattr(self, "confidence_var") else "Not calculated"
        missing_count = len(evidence.get("missing_critical", []))
        document_state = "Not generated"
        if hasattr(self, "documentation_text") and self.documentation_text.get("1.0", "end").strip():
            document_state = self.document_type_var.get().strip() or "Draft ready"

        self.patient_kpi_vars["risk"].set(current_risk or "Not calculated")
        self.patient_kpi_vars["pathway"].set(current_pathway or "Not calculated")
        self.patient_kpi_vars["confidence"].set(current_confidence or "Not calculated")
        self.patient_kpi_vars["missing"].set(str(missing_count))
        self.patient_kpi_vars["document"].set(document_state)

        risk_color = {"High Risk": "#b91c1c", "Medium Risk": "#b45309", "Low Risk": "#166534"}.get(current_risk, UI_COLORS["text"])
        conf_color = {"High Confidence": "#166534", "Moderate Confidence": "#b45309", "Low Confidence": "#b91c1c"}.get(current_confidence, UI_COLORS["text"])
        if hasattr(self, "patient_kpi_value_labels"):
            self.patient_kpi_value_labels["risk"].configure(fg=risk_color)
            self.patient_kpi_value_labels["confidence"].configure(fg=conf_color)
            self.patient_kpi_value_labels["missing"].configure(fg="#b91c1c" if missing_count else "#166534")

    def update_audit_summary_cards(self):
        if not hasattr(self, "audit_kpi_vars"):
            return
        rows = self.load_audit_rows()
        self.audit_kpi_vars["total"].set(str(len(rows)))
        self.audit_kpi_vars["overrides"].set(str(sum(1 for r in rows if r.get("event_type") == "CLINICIAN_OVERRIDE_APPLIED")))
        self.audit_kpi_vars["documents"].set(str(sum(1 for r in rows if r.get("event_type") == "DOCUMENT_GENERATED")))
        self.audit_kpi_vars["confirmed"].set(str(sum(1 for r in rows if r.get("event_type") == "FINAL_DECISION_CONFIRMED")))
        self.audit_kpi_vars["high_risk"].set(str(sum(1 for r in rows if r.get("risk_output") == "High Risk")))

    def ensure_audit_file(self):
        try:
            file_needs_header = True

            if os.path.exists(AUDIT_FILE) and os.path.getsize(AUDIT_FILE) > 0:
                with open(AUDIT_FILE, "r", encoding="utf-8") as file:
                    first_line = file.readline().strip()

                expected_start = "event_id,timestamp,event_type"
                file_needs_header = not first_line.startswith(expected_start)

                if file_needs_header:
                    backup_file = AUDIT_FILE.replace(
                        ".csv",
                        f"_BAD_HEADER_BACKUP_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
                    )
                    os.replace(AUDIT_FILE, backup_file)

            if file_needs_header:
                with open(AUDIT_FILE, "w", newline="", encoding="utf-8") as file:
                    writer = csv.DictWriter(file, fieldnames=AUDIT_FIELDNAMES)
                    writer.writeheader()

        except Exception as error:
            messagebox.showerror(
                "Audit Log Error",
                f"Could not prepare audit log file.\n\nError:\n{error}"
            )

    def load_audit_rows(self):
        self.ensure_audit_file()
        try:
            with open(AUDIT_FILE, "r", newline="", encoding="utf-8") as file:
                return list(csv.DictReader(file))
        except Exception:
            return []

    def append_audit_log(self, event_type, clinician_action="", notes="", document_type="", document_status="Draft generated"):
        global AUDIT_FILE
        self.ensure_audit_file()
        timestamp = datetime.now()
        patient_id = self.case_vars["patient_id"].get().strip() if hasattr(self, "case_vars") else ""
        patient_name = self.case_vars["patient_name"].get().strip() if hasattr(self, "case_vars") else ""
        user_role = self.user_role_var.get().strip() if hasattr(self, "user_role_var") else "Clinician"
        risk_output = self.output_risk_var.get().strip() if hasattr(self, "output_risk_var") else ""
        risk_score = self.risk_score_var.get().strip() if hasattr(self, "risk_score_var") else ""
        final_pathway = self.output_pathway_var.get().strip() if hasattr(self, "output_pathway_var") else ""
        recommended_action = self.recommended_action_var.get().strip() if hasattr(self, "recommended_action_var") else ""
        confidence_level = self.confidence_var.get().strip() if hasattr(self, "confidence_var") else ""
        trigger_summary = ""
        if hasattr(self, "case_vars") and "trigger_summary" in self.case_vars:
            trigger_summary = self.case_vars["trigger_summary"].get().strip()
        if not trigger_summary and hasattr(self, "case_evidence_text"):
            trigger_summary = (self.case_evidence_text.get("1.0", "end").strip().splitlines() or [""])[0]

        row = {
            "event_id": timestamp.strftime("%Y%m%d%H%M%S%f"),
            "timestamp": timestamp.isoformat(timespec="seconds"),
            "event_type": event_type,
            "user_role": user_role,
            "patient_id": patient_id,
            "patient_name": patient_name,
            "target_output": self.get_active_prediction_target() if hasattr(self, "get_active_prediction_target") else "",
            "risk_output": risk_output,
            "risk_score": risk_score,
            "final_pathway": final_pathway,
            "recommended_action": recommended_action,
            "clinician_action": clinician_action,
            "override_risk": self.override_risk_var.get().strip() if hasattr(self, "override_risk_var") else "",
            "override_pathway": self.override_pathway_var.get().strip() if hasattr(self, "override_pathway_var") else "",
            "override_reason": self.override_reason_text.get("1.0", "end").strip() if hasattr(self, "override_reason_text") else "",
            "document_type": document_type or (self.document_type_var.get().strip() if hasattr(self, "document_type_var") else ""),
            "document_status": document_status if document_type or (hasattr(self, "document_type_var") and self.document_type_var.get().strip()) else "",
            "confidence_level": confidence_level,
            "trigger_summary": trigger_summary,
            "outcome_status": final_pathway,
            "notes": notes,
            "local_ed_acuity": self.local_acuity_var.get().strip() if hasattr(self, "local_acuity_var") else "",
            "safety_lock_status": self.safety_lock_var.get().strip() if hasattr(self, "safety_lock_var") else "",
            "clinician_confirmation_status": self.clinician_confirmation_var.get().strip() if hasattr(self, "clinician_confirmation_var") else "",
            "emergency_gates": self.triage_gate_var.get().strip() if hasattr(self, "triage_gate_var") else "",
        }
        try:
            with open(AUDIT_FILE, "a", newline="", encoding="utf-8") as file:
                writer = csv.DictWriter(file, fieldnames=AUDIT_FIELDNAMES)
                writer.writerow(row)
        except Exception as error:
            # If Excel/Windows locks the main CSV, automatically switch to a writable backup
            # instead of interrupting the clinical workflow.
            try:
                backup_file = _resolve_runtime_data_file("hf_triage_audit_log_backup.csv")
                file_exists = os.path.exists(backup_file) and os.path.getsize(backup_file) > 0
                with open(backup_file, "a", newline="", encoding="utf-8") as file:
                    writer = csv.DictWriter(file, fieldnames=AUDIT_FIELDNAMES)
                    if not file_exists:
                        writer.writeheader()
                    writer.writerow(row)
                AUDIT_FILE = backup_file
                messagebox.showwarning(
                    "Audit Log Switched to Backup",
                    "The main audit CSV was locked or not writable, so the program switched to a backup audit log.\n\n"
                    f"Backup file:\n{backup_file}\n\nOriginal error:\n{error}",
                )
            except Exception as backup_error:
                messagebox.showerror("Audit Log Error", f"Could not append audit log.\n\n{error}\n\nBackup also failed:\n{backup_error}")
                return
        

                
        if hasattr(self, "audit_tree"):
            self.refresh_audit_tree()
        else:
            self.update_audit_summary_cards()

    def get_current_patient_key(self):
        patient_id = self.case_vars["patient_id"].get().strip() if hasattr(self, "case_vars") else ""
        patient_name = self.case_vars["patient_name"].get().strip() if hasattr(self, "case_vars") else ""
        if patient_id:
            return patient_id
        if patient_name:
            return f"NAME::{patient_name.lower()}"
        return "UNSAVED_CURRENT_CASE"

    def get_alert_delivery_mode(self, risk_output):
        return ALERT_DELIVERY_RULES.get(risk_output, "Not calculated")

    def apply_alert_fatigue_control(self, evidence):
        risk_output = evidence.get("risk_output", "")
        patient_key = self.get_current_patient_key()
        trigger_signature = " | ".join(evidence.get("criteria_used", [])[:6]) or evidence.get("risk_output", "")
        now = datetime.now()
        state = self.alert_state_by_patient.get(patient_key, {})
        delivery_mode = self.get_alert_delivery_mode(risk_output)
        suppression_active = False
        fatigue_status = "No active alert fatigue rule applied yet"
        interruptive_active = False

        if risk_output == "High Risk":
            interruptive_active = True
            acknowledged_at = state.get("acknowledged_at")
            same_signature = state.get("trigger_signature") == trigger_signature
            same_risk = state.get("risk_output") == risk_output
            if acknowledged_at and same_signature and same_risk:
                elapsed = now - acknowledged_at
                if elapsed < timedelta(minutes=ALERT_SUPPRESSION_WINDOW_MINUTES):
                    remaining_minutes = max(1, int((timedelta(minutes=ALERT_SUPPRESSION_WINDOW_MINUTES) - elapsed).total_seconds() // 60))
                    suppression_active = True
                    interruptive_active = False
                    delivery_mode = "Interruptive alert suppressed; dashboard warning retained"
                    evidence["early_warning_alert"] = "High-risk condition still present - duplicate interruptive alert suppressed after recent acknowledgement"
                    fatigue_status = f"Duplicate high-risk alert suppressed for {remaining_minutes} more minute(s) after acknowledgement"
                    evidence["recommended_action"] = evidence.get("recommended_action", "") + " Review the already acknowledged high-risk alert unless the clinical picture worsens."
                else:
                    fatigue_status = "New interruptive high-risk alert allowed because the suppression window expired"
            else:
                fatigue_status = "Interruptive alert active for a new or changed high-risk pattern"
        elif risk_output == "Medium Risk":
            interruptive_active = False
            delivery_mode = "Non-interruptive dashboard flag only"
            fatigue_status = "Medium-risk output is shown as a dashboard flag to reduce alert fatigue"
        elif risk_output == "Low Risk":
            interruptive_active = False
            delivery_mode = "Silent monitoring only"
            fatigue_status = "Low-risk output does not create an interruptive alert"
        elif risk_output.startswith("Blocked"):
            interruptive_active = False
            delivery_mode = "Safety block - final triage disabled"
            fatigue_status = "No alert fatigue rule applied because the case is blocked by safety validation"

        evidence["alert_delivery_mode"] = delivery_mode
        evidence["alert_fatigue_status"] = fatigue_status
        evidence["interruptive_alert_active"] = interruptive_active
        evidence["suppression_active"] = suppression_active
        evidence["patient_alert_key"] = patient_key
        evidence["trigger_signature"] = trigger_signature
        self.current_alert_context = {
            "patient_key": patient_key,
            "risk_output": risk_output,
            "trigger_signature": trigger_signature,
            "delivery_mode": delivery_mode,
            "suppression_active": suppression_active,
            "interruptive_alert_active": interruptive_active,
            "fatigue_status": fatigue_status,
        }
        return evidence

    def acknowledge_current_alert(self):
        context = dict(self.current_alert_context) if hasattr(self, "current_alert_context") else {}
        risk_output = self.output_risk_var.get().strip() if hasattr(self, "output_risk_var") else ""
        if not risk_output or risk_output == "Not calculated":
            messagebox.showwarning("Alert Acknowledgement", "Run the triage result first before acknowledging an alert.")
            return
        if risk_output == "Low Risk":
            messagebox.showinfo("Alert Acknowledgement", "Low-risk cases do not create an interruptive alert to acknowledge.")
            return
        patient_key = context.get("patient_key") or self.get_current_patient_key()
        self.alert_state_by_patient[patient_key] = {
            "acknowledged_at": datetime.now(),
            "risk_output": context.get("risk_output", risk_output),
            "trigger_signature": context.get("trigger_signature", ""),
        }
        if risk_output == "High Risk":
            self.alert_fatigue_var.set(f"High-risk alert acknowledged. Duplicate interruptive alerts will be suppressed for {ALERT_SUPPRESSION_WINDOW_MINUTES} minutes unless the pattern changes.")
        else:
            self.alert_fatigue_var.set("Medium-risk dashboard flag acknowledged. No interruptive alert will repeat for this unchanged case state.")
        self.update_status("Current alert acknowledged for fatigue control")

    def reset_alert_suppression(self):
        patient_key = self.get_current_patient_key()
        if patient_key in self.alert_state_by_patient:
            del self.alert_state_by_patient[patient_key]
            self.alert_fatigue_var.set("Alert suppression reset for the current patient. The next qualifying high-risk result can trigger a fresh interruptive alert.")
            self.update_status("Alert suppression reset")
        else:
            messagebox.showinfo("Alert Suppression", "No alert suppression record exists for the current patient.")

    def build_audit_tab(self):
        audit_banner = self.create_banner(
            self.audit_tab,
            "Audit Trail & Review Monitoring",
            "Use this tab to inspect every AI result, generated document, clinician override, and final decision confirmation recorded by the prototype.",
            bg="#102a43",
        )
        audit_banner.pack(fill="x", pady=(0, 10))

        audit_cards_row = ttk.Frame(self.audit_tab)
        audit_cards_row.pack(fill="x", pady=(0, 10))
        self.audit_kpi_vars = {
            "total": tk.StringVar(value="0"),
            "overrides": tk.StringVar(value="0"),
            "documents": tk.StringVar(value="0"),
            "confirmed": tk.StringVar(value="0"),
            "high_risk": tk.StringVar(value="0"),
        }
        audit_card_specs = [
            ("Total Events", "total", "All audit entries in the current CSV", UI_COLORS["blue_soft"]),
            ("Overrides", "overrides", "Clinician override actions", UI_COLORS["red_soft"]),
            ("Documents", "documents", "Generated structured notes", UI_COLORS["teal_soft"]),
            ("Confirmed", "confirmed", "Final decision saves", UI_COLORS["green_soft"]),
            ("High-Risk Events", "high_risk", "Rows currently tagged High Risk", UI_COLORS["amber_soft"]),
        ]
        self.audit_kpi_value_labels = {}
        for title, key, subtitle, bg in audit_card_specs:
            card, value_label = self.create_metric_card(audit_cards_row, title, self.audit_kpi_vars[key], subtitle=subtitle, bg=bg, width=218)
            card.pack(side="left", fill="x", expand=True, padx=(0, 8))
            self.audit_kpi_value_labels[key] = value_label

        filter_frame = ttk.LabelFrame(self.audit_tab, text="Audit Filters", padding=8)
        filter_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(filter_frame, text="Search:").pack(side="left", padx=(0, 5))
        self.audit_search_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=self.audit_search_var, width=32, style="Compact.TEntry").pack(side="left")
        self.audit_search_var.trace_add("write", lambda *args: self.refresh_audit_tree())

        ttk.Label(filter_frame, text="Event Type:").pack(side="left", padx=(12, 5))
        self.audit_event_filter_var = tk.StringVar(value="All")
        ttk.Combobox(
            filter_frame,
            textvariable=self.audit_event_filter_var,
            values=["All", "PATIENT_REGISTERED", "AI_RESULT_DISPLAYED", "DOCUMENT_GENERATED", "CLINICIAN_OVERRIDE_APPLIED", "FINAL_DECISION_CONFIRMED"],
            state="readonly",
            width=24,
            style="Compact.TCombobox",
        ).pack(side="left")
        self.audit_event_filter_var.trace_add("write", lambda *args: self.refresh_audit_tree())

        ttk.Button(filter_frame, text="Refresh", style="Compact.TButton", command=self.refresh_audit_tree).pack(side="right", padx=4)
        ttk.Button(filter_frame, text="Export Audit CSV", style="Compact.TButton", command=self.export_audit_log_csv).pack(side="right", padx=4)

        ttk.Label(filter_frame, text="Tip: search by patient ID, event type, role, pathway, or override reason.", style="SectionHint.TLabel").pack(side="right", padx=8)

        tree_frame = ttk.LabelFrame(self.audit_tab, text="AI Decision Audit Trail", padding=8)
        tree_frame.pack(fill="both", expand=True)

        audit_columns = (
            "timestamp", "patient_id", "patient_name", "event_type", "user_role",
            "risk_output", "final_pathway", "clinician_action", "document_type", "override_reason"
        )
        self.audit_tree = ttk.Treeview(tree_frame, columns=audit_columns, show="headings")
        headings = {
            "timestamp": "Timestamp",
            "patient_id": "Patient ID",
            "patient_name": "Patient Name",
            "event_type": "Event Type",
            "user_role": "User Role",
            "risk_output": "Risk Output",
            "final_pathway": "Final Pathway",
            "clinician_action": "Clinician Action",
            "document_type": "Document Type",
            "override_reason": "Override Reason",
        }
        widths = {
            "timestamp": 145, "patient_id": 100, "patient_name": 135, "event_type": 180,
            "user_role": 120, "risk_output": 95, "final_pathway": 110, "clinician_action": 120,
            "document_type": 115, "override_reason": 280,
        }
        for column in audit_columns:
            self.audit_tree.heading(column, text=headings[column])
            self.audit_tree.column(column, width=widths[column], anchor="w")
        self.audit_tree.pack(side="left", fill="both", expand=True)
        self.audit_tree.tag_configure("audit_override", background="#fff1f2")
        self.audit_tree.tag_configure("audit_document", background="#eff6ff")
        self.audit_tree.tag_configure("audit_confirmed", background="#ecfdf3")
        self.audit_tree.tag_configure("audit_result", background="#fffbeb")
        self.audit_tree.tag_configure("audit_registered", background="#f8fafc")

        audit_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.audit_tree.yview)
        audit_scroll.pack(side="right", fill="y")
        self.audit_tree.configure(yscrollcommand=audit_scroll.set)

        # Load and display existing audit rows immediately when the tab is built.
        self.refresh_audit_tree()

    def refresh_audit_tree(self):
        if not hasattr(self, "audit_tree"):
            return
        for item in self.audit_tree.get_children():
            self.audit_tree.delete(item)

        search_text = self.audit_search_var.get().strip().lower() if hasattr(self, "audit_search_var") else ""
        event_filter = self.audit_event_filter_var.get().strip() if hasattr(self, "audit_event_filter_var") else "All"

        rows = self.load_audit_rows()
        for row in reversed(rows):
            haystack = " ".join(str(row.get(key, "")) for key in AUDIT_FIELDNAMES).lower()
            if search_text and search_text not in haystack:
                continue
            if event_filter and event_filter != "All" and row.get("event_type", "") != event_filter:
                continue
            tag = ()
            event_type = row.get("event_type", "")
            if event_type == "CLINICIAN_OVERRIDE_APPLIED":
                tag = ("audit_override",)
            elif event_type == "DOCUMENT_GENERATED":
                tag = ("audit_document",)
            elif event_type == "FINAL_DECISION_CONFIRMED":
                tag = ("audit_confirmed",)
            elif event_type == "AI_RESULT_DISPLAYED":
                tag = ("audit_result",)
            elif event_type == "PATIENT_REGISTERED":
                tag = ("audit_registered",)
            self.audit_tree.insert("", "end", values=(
                row.get("timestamp", ""),
                row.get("patient_id", ""),
                row.get("patient_name", ""),
                row.get("event_type", ""),
                row.get("user_role", ""),
                row.get("risk_output", ""),
                row.get("final_pathway", ""),
                row.get("clinician_action", ""),
                row.get("document_type", ""),
                row.get("override_reason", ""),
            ), tags=tag)
        self.update_audit_summary_cards()

    def export_audit_log_csv(self):
        rows = self.load_audit_rows()
        if not rows:
            messagebox.showwarning("Export", "No audit log rows to export.")
            return
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="Export Audit Log CSV",
        )
        if not filename:
            return
        try:
            with open(filename, "w", newline="", encoding="utf-8") as file:
                writer = csv.DictWriter(file, fieldnames=AUDIT_FIELDNAMES)
                writer.writeheader()
                writer.writerows(rows)
            messagebox.showinfo("Exported", "Audit log CSV exported successfully.")
            self.update_status("Audit log CSV exported")
        except Exception as error:
            messagebox.showerror("Export Error", f"Could not export audit log CSV.\n\n{error}")

    def get_active_prediction_target(self):
        if hasattr(self, "case_target_var"):
            target = self.case_target_var.get().strip()
            if target:
                return target
        return self.project_target_var.get()

    def get_predictive_use_case_metadata(self, target=None):
        target = target or self.get_active_prediction_target()
        return {
            "target_output": target,
            "prediction_timing": PREDICTION_TIMING_TEXT,
            "target_users": PREDICTION_USERS_TEXT,
            "output_type": PREDICTION_OUTPUT_TEXT,
            "clinician_oversight": CLINICIAN_OVERSIGHT_TEXT,
        }

    def on_case_target_change(self, event=None):
        self.update_status(f"Prediction target set to: {self.get_active_prediction_target()}")
        has_case_data = hasattr(self, "case_vars") and (self.case_vars["patient_id"].get().strip() or self.case_vars["patient_name"].get().strip())
        if has_case_data:
            self.generate_case_summary()

    def save_project_target(self):
        self.app_data["project_target"] = self.project_target_var.get()
        if hasattr(self, "case_target_var") and self.selected_case_index is None and not self.case_target_var.get().strip():
            self.case_target_var.set(self.project_target_var.get())
        self.storage.save_data(self.app_data)
        self.update_status("Project target saved")

    def save_all_data(self):
        self.app_data["project_target"] = self.project_target_var.get()
        self.app_data["workflow_steps"] = self.workflow_steps
        self.app_data["data_inventory"] = self.data_inventory
        self.app_data["patient_cases"] = self.patient_cases
        try:
            self.storage.save_data(self.app_data)
            messagebox.showinfo("Saved", "All data saved locally.")
            self.update_status("All data saved locally")
        except Exception as error:
            messagebox.showerror("Save Error", f"Could not save data.\n\n{error}")

    def reload_local_data(self):
        self.app_data = self.storage.load_data()
        self.project_target_var.set(self.app_data.get("project_target", PREDICTION_TARGETS[1]))
        self.workflow_steps = self.app_data["workflow_steps"]
        self.data_inventory = self.app_data["data_inventory"]
        self.patient_cases = self.app_data["patient_cases"]
        self.refresh_workflow_tree()
        self.refresh_data_tree()
        self.refresh_case_tree()
        self.clear_workflow_form()
        self.clear_data_form()
        self.clear_case_form()
        self.update_status("Data reloaded from local storage")

    def clear_saved_patient_cases(self):
        if not messagebox.askyesno(
            "Clear Saved Cases",
            "Delete the actual local JSON file and rebuild it without saved patient cases?\n\nThis keeps workflow steps and data inventory, but removes all imported/saved patient cases from the real storage file."
        ):
            return

        preserved_data = {
            "project_target": self.project_target_var.get(),
            "workflow_steps": self.workflow_steps,
            "data_inventory": self.data_inventory,
            "patient_cases": [],
        }

        try:
            self.storage.delete_file()
            self.storage.save_data(preserved_data)
            self.app_data = preserved_data
            self.patient_cases = []
            self.refresh_case_tree()
            self.clear_case_form()
            self.update_status(f"Saved patient cases cleared from actual file: {self.storage.filename}")
            messagebox.showinfo(
                "Cleared",
                f"The actual local JSON file was reset successfully.\n\nFile: {self.storage.filename}\n\nYou can now import your CSV again."
            )
        except Exception as error:
            messagebox.showerror("Clear Error", f"Could not reset the actual local JSON file.\n\n{error}")

    # ------------------------- Phase 1 -------------------------
    def build_phase1_tab(self):
        info_frame = ttk.LabelFrame(self.phase1_tab, text="Presentation Workflow Alignment", padding=8)
        info_frame.pack(fill="x", pady=(0, 8))
        ttk.Label(
            info_frame,
            text="This tab uses the workflow presets aligned with your presentation. Load the current or AI workflow below and edit the steps directly without showing a large preview area.",
            wraplength=1400,
            justify="left"
        ).pack(anchor="w")

        self.workflow_map_image = None

        form_frame = ttk.LabelFrame(self.phase1_tab, text="Workflow Step Form", padding=8)
        form_frame.pack(fill="x", pady=(0, 8))

        self.workflow_vars = {
            "step_number": tk.StringVar(),
            "step_name": tk.StringVar(),
            "actor": tk.StringVar(),
            "input": tk.StringVar(),
            "action": tk.StringVar(),
            "output": tk.StringVar(),
        }

        labels = [
            ("Step Number", "step_number", 0, 0),
            ("Step Name", "step_name", 0, 2),
            ("Actor", "actor", 1, 0),
            ("Input", "input", 1, 2),
            ("Action", "action", 2, 0),
            ("Output", "output", 2, 2),
        ]

        for text_label, key, row, col in labels:
            ttk.Label(form_frame, text=text_label).grid(row=row, column=col, sticky="w", padx=5, pady=5)
            if key == "step_name":
                combo = ttk.Combobox(form_frame, textvariable=self.workflow_vars[key], values=list(WORKFLOW_SUGGESTIONS.keys()), width=28)
                combo.grid(row=row, column=col + 1, sticky="ew", padx=5, pady=5)
            else:
                ttk.Entry(form_frame, textvariable=self.workflow_vars[key], width=44).grid(row=row, column=col + 1, sticky="ew", padx=5, pady=5)

        for column in range(4):
            form_frame.columnconfigure(column, weight=1)

        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=3, column=0, columnspan=4, sticky="ew", pady=(10, 0))

        ttk.Button(button_frame, text="Apply Suggested Step", command=self.apply_workflow_suggestion).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Add Step", command=self.add_workflow_step).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Update Step", command=self.update_workflow_step).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Delete Step", command=self.delete_workflow_step).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Move Up", command=lambda: self.move_workflow_step(-1)).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Move Down", command=lambda: self.move_workflow_step(1)).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Reset Form", style="Compact.TButton", command=self.clear_workflow_form).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Load Current Workflow", command=self.load_default_workflow).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Load AI Workflow", command=self.load_ai_workflow).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Check Workflow Coverage", command=self.check_workflow_coverage).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Save", style="Compact.TButton", command=self.save_all_data).pack(side="right", padx=4)
        ttk.Button(button_frame, text="Export CSV", style="Compact.TButton", command=self.export_workflow_csv).pack(side="right", padx=4)

        tree_frame = ttk.LabelFrame(self.phase1_tab, text="Workflow Steps (Exact Presentation Match)", padding=8)
        tree_frame.pack(fill="both", expand=True)
        self._bind_scroll_to_children(self.patient_tab)

        workflow_columns = ("step_number", "step_name", "actor", "input", "action", "output")
        self.workflow_tree = ttk.Treeview(tree_frame, columns=workflow_columns, show="headings")
        for column in workflow_columns:
            self.workflow_tree.heading(column, text=column.replace("_", " ").title())
        self.workflow_tree.column("step_number", width=90, anchor="center")
        self.workflow_tree.column("step_name", width=220)
        self.workflow_tree.column("actor", width=210)
        self.workflow_tree.column("input", width=260)
        self.workflow_tree.column("action", width=270)
        self.workflow_tree.column("output", width=240)
        self.workflow_tree.pack(side="left", fill="both", expand=True)
        self.workflow_tree.bind("<<TreeviewSelect>>", self.on_workflow_select)

        workflow_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.workflow_tree.yview)
        workflow_scroll.pack(side="right", fill="y")
        self.workflow_tree.configure(yscrollcommand=workflow_scroll.set)

    def apply_workflow_suggestion(self):
        step_name = self.workflow_vars["step_name"].get().strip()
        if step_name not in WORKFLOW_SUGGESTIONS:
            messagebox.showwarning("Suggestion", "Choose a suggested step name first.")
            return
        suggestion = WORKFLOW_SUGGESTIONS[step_name]
        self.workflow_vars["actor"].set(suggestion["actor"])
        self.workflow_vars["input"].set(suggestion["input"])
        self.workflow_vars["action"].set(suggestion["action"])
        self.workflow_vars["output"].set(suggestion["output"])
        if not self.workflow_vars["step_number"].get().strip():
            self.workflow_vars["step_number"].set(str(len(self.workflow_steps) + 1))
        self.update_status("Workflow suggestion applied")

    def validate_workflow_inputs(self, for_update=False):
        values = {key: var.get().strip() for key, var in self.workflow_vars.items()}
        required_fields = ["step_number", "step_name", "actor", "action"]
        for field in required_fields:
            if not values[field]:
                messagebox.showwarning("Validation", f"{field.replace('_', ' ').title()} is required.")
                return None
        try:
            step_number = int(values["step_number"])
            if step_number <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Validation", "Step Number must be a positive integer.")
            return None

        for index, step in enumerate(self.workflow_steps):
            if step["step_number"] == step_number:
                if not for_update or index != self.selected_workflow_index:
                    messagebox.showwarning("Validation", "Duplicate Step Number is not allowed.")
                    return None
        values["step_number"] = step_number
        return values

    def add_workflow_step(self):
        values = self.validate_workflow_inputs(for_update=False)
        if not values:
            return
        self.workflow_steps.append(values)
        self.sort_workflow_steps()
        self.refresh_workflow_tree()
        self.clear_workflow_form()
        self.update_status("Workflow step added")
        messagebox.showinfo("Added", "Workflow step added successfully.")

    def update_workflow_step(self):
        if self.selected_workflow_index is None:
            messagebox.showwarning("Update", "Select a workflow step to update.")
            return
        values = self.validate_workflow_inputs(for_update=True)
        if not values:
            return
        self.workflow_steps[self.selected_workflow_index] = values
        self.sort_workflow_steps()
        self.refresh_workflow_tree()
        self.clear_workflow_form()
        self.update_status("Workflow step updated")
        messagebox.showinfo("Updated", "Workflow step updated successfully.")

    def delete_workflow_step(self):
        if self.selected_workflow_index is None:
            messagebox.showwarning("Delete", "Select a workflow step to delete.")
            return
        if messagebox.askyesno("Confirm Delete", "Delete the selected workflow step?"):
            del self.workflow_steps[self.selected_workflow_index]
            self.refresh_workflow_tree()
            self.clear_workflow_form()
            self.update_status("Workflow step deleted")

    def move_workflow_step(self, direction):
        if self.selected_workflow_index is None:
            messagebox.showwarning("Move", "Select a workflow step first.")
            return
        new_index = self.selected_workflow_index + direction
        if new_index < 0 or new_index >= len(self.workflow_steps):
            return
        self.workflow_steps[self.selected_workflow_index], self.workflow_steps[new_index] = self.workflow_steps[new_index], self.workflow_steps[self.selected_workflow_index]
        self.reassign_workflow_numbers()
        self.refresh_workflow_tree(select_index=new_index)
        self.update_status("Workflow step reordered")

    def reassign_workflow_numbers(self):
        for index, step in enumerate(self.workflow_steps, start=1):
            step["step_number"] = index

    def sort_workflow_steps(self):
        self.workflow_steps.sort(key=lambda item: item["step_number"])

    def refresh_workflow_tree(self, select_index=None):
        for item in self.workflow_tree.get_children():
            self.workflow_tree.delete(item)
        self.sort_workflow_steps()
        for index, step in enumerate(self.workflow_steps):
            item_id = self.workflow_tree.insert("", "end", values=(
                step["step_number"], step["step_name"], step["actor"], step["input"], step["action"], step["output"]
            ))
            if select_index is not None and index == select_index:
                self.workflow_tree.selection_set(item_id)
                self.workflow_tree.focus(item_id)
                self.selected_workflow_index = select_index

    def on_workflow_select(self, event=None):
        selection = self.workflow_tree.selection()
        if not selection:
            return
        item_id = selection[0]
        values = self.workflow_tree.item(item_id, "values")
        children = self.workflow_tree.get_children()
        self.selected_workflow_index = children.index(item_id)
        keys = ["step_number", "step_name", "actor", "input", "action", "output"]
        for key, value in zip(keys, values):
            self.workflow_vars[key].set(value)

    def clear_workflow_form(self):
        for var in self.workflow_vars.values():
            var.set("")
        self.selected_workflow_index = None

    def load_default_workflow(self):
        if messagebox.askyesno("Load Workflow", "Replace current workflow with the exact current clinical workflow from the presentation?"):
            self.workflow_steps = default_workflow_steps("current")
            self.refresh_workflow_tree()
            self.clear_workflow_form()
            self.update_status("Exact current clinical workflow loaded")

    def load_ai_workflow(self):
        if messagebox.askyesno("Load AI Workflow", "Replace current workflow with the exact AI-assisted workflow from the presentation?"):
            self.workflow_steps = default_workflow_steps("ai")
            self.refresh_workflow_tree()
            self.clear_workflow_form()
            self.update_status("Exact AI-assisted workflow loaded")

    def check_workflow_coverage(self):
        existing_steps = {step["step_name"].strip().lower() for step in self.workflow_steps}
        missing = [name for name in REQUIRED_WORKFLOW_STEPS if name.strip().lower() not in existing_steps]
        if missing:
            messagebox.showwarning("Workflow Coverage", "Missing workflow steps:\n\n- " + "\n- ".join(missing))
        else:
            messagebox.showinfo("Workflow Coverage", "All current-clinical workflow steps from the presentation are present.")
        self.update_status("Current-clinical workflow coverage checked")

    def export_workflow_csv(self):
        if not self.workflow_steps:
            messagebox.showwarning("Export", "No workflow data to export.")
            return
        filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")], title="Export Workflow CSV")
        if not filename:
            return
        try:
            with open(filename, "w", newline="", encoding="utf-8") as file:
                writer = csv.DictWriter(file, fieldnames=["step_number", "step_name", "actor", "input", "action", "output"])
                writer.writeheader()
                writer.writerows(self.workflow_steps)
            messagebox.showinfo("Exported", "Workflow CSV exported successfully.")
            self.update_status("Workflow CSV exported")
        except Exception as error:
            messagebox.showerror("Export Error", f"Could not export workflow CSV.\n\n{error}")

    # ------------------------- Phase 2 -------------------------
    def build_phase2_tab(self):
        search_frame = ttk.Frame(self.phase2_tab)
        search_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(search_frame, text="Search:").pack(side="left", padx=(0, 5))
        self.data_search_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.data_search_var, width=28).pack(side="left")
        self.data_search_var.trace_add("write", lambda *args: self.refresh_data_tree())

        ttk.Label(search_frame, text="Category Filter:").pack(side="left", padx=(15, 5))
        self.data_filter_var = tk.StringVar(value="All")
        filter_values = ["All", "Demographics", "Symptoms", "Vital signs", "Medical history", "Red flags", "Clinical notes"]
        filter_combo = ttk.Combobox(search_frame, textvariable=self.data_filter_var, values=filter_values, state="readonly", width=18)
        filter_combo.pack(side="left")
        filter_combo.bind("<<ComboboxSelected>>", lambda event: self.refresh_data_tree())

        form_frame = ttk.LabelFrame(self.phase2_tab, text="Data Inventory Form", padding=8)
        form_frame.pack(fill="x", pady=(0, 8))

        self.data_vars = {
            "category": tk.StringVar(),
            "data_element": tk.StringVar(),
            "data_type": tk.StringVar(),
            "source": tk.StringVar(),
            "why_it_matters": tk.StringVar(),
            "priority": tk.StringVar(),
            "available_at_triage": tk.StringVar(),
            "status": tk.StringVar(),
        }

        field_specs = [
            ("Category", "category", ["Demographics", "Symptoms", "Vital signs", "Medical history", "Red flags", "Clinical notes"]),
            ("Data Element", "data_element", list(DATA_ELEMENT_SUGGESTIONS.keys())),
            ("Data Type", "data_type", ["Structured", "Semi-structured", "Unstructured"]),
            ("Source", "source", ["Registration", "Triage form / nurse interview", "Triage vital signs", "EMR", "EMR / patient interview", "Nurse observation", "Clinical assessment", "Nurse note", "Referral document", "Pulse oximeter", "Blood pressure", "Patient interview / triage note"]),
            ("Why It Matters", "why_it_matters", None),
            ("Priority", "priority", ["High", "Medium", "Low"]),
            ("Available at Triage?", "available_at_triage", ["Yes", "No", "Sometimes"]),
            ("Status", "status", ["Keep", "Add", "Exclude"]),
        ]

        for index, (label_text, key, values) in enumerate(field_specs):
            row = index // 4
            col = (index % 4) * 2
            ttk.Label(form_frame, text=label_text).grid(row=row, column=col, sticky="w", padx=5, pady=5)
            if values is None:
                ttk.Entry(form_frame, textvariable=self.data_vars[key], width=28).grid(row=row, column=col + 1, sticky="ew", padx=5, pady=5)
            else:
                combo = ttk.Combobox(form_frame, textvariable=self.data_vars[key], values=values, width=27)
                combo.grid(row=row, column=col + 1, sticky="ew", padx=5, pady=5)

        for column in range(8):
            form_frame.columnconfigure(column, weight=1)

        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=3, column=0, columnspan=8, sticky="ew", pady=(10, 0))

        ttk.Button(button_frame, text="Apply Suggested Data Element", command=self.apply_data_suggestion).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Add Element", command=self.add_data_item).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Update Element", command=self.update_data_item).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Delete Element", command=self.delete_data_item).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Reset Form", style="Compact.TButton", command=self.clear_data_form).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Load Full Project Data", command=self.load_default_data_inventory).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Check Project Coverage", command=self.check_data_coverage).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Save", style="Compact.TButton", command=self.save_all_data).pack(side="right", padx=4)
        ttk.Button(button_frame, text="Export CSV", style="Compact.TButton", command=self.export_data_csv).pack(side="right", padx=4)

        tree_frame = ttk.LabelFrame(self.phase2_tab, text="Data Inventory", padding=8)
        tree_frame.pack(fill="both", expand=True)
        self._bind_scroll_to_children(self.patient_tab)

        data_columns = ("category", "data_element", "data_type", "source", "why_it_matters", "priority", "available_at_triage", "status")
        self.data_tree = ttk.Treeview(tree_frame, columns=data_columns, show="headings")
        for column in data_columns:
            self.data_tree.heading(column, text=column.replace("_", " ").title())
        self.data_tree.column("category", width=120)
        self.data_tree.column("data_element", width=170)
        self.data_tree.column("data_type", width=120)
        self.data_tree.column("source", width=190)
        self.data_tree.column("why_it_matters", width=260)
        self.data_tree.column("priority", width=80, anchor="center")
        self.data_tree.column("available_at_triage", width=120, anchor="center")
        self.data_tree.column("status", width=90, anchor="center")
        self.data_tree.pack(side="left", fill="both", expand=True)
        self.data_tree.bind("<<TreeviewSelect>>", self.on_data_select)
        self.data_tree.tag_configure("priority_attention", background="#fff1b8")

        data_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.data_tree.yview)
        data_scroll.pack(side="right", fill="y")
        self.data_tree.configure(yscrollcommand=data_scroll.set)

    def apply_data_suggestion(self):
        element_name = self.data_vars["data_element"].get().strip()
        if element_name not in DATA_ELEMENT_SUGGESTIONS:
            messagebox.showwarning("Suggestion", "Choose a suggested data element first.")
            return
        suggestion = DATA_ELEMENT_SUGGESTIONS[element_name]
        self.data_vars["category"].set(suggestion["category"])
        self.data_vars["data_type"].set(suggestion["data_type"])
        self.data_vars["source"].set(suggestion["source"])
        self.data_vars["why_it_matters"].set(suggestion["why_it_matters"])
        self.data_vars["priority"].set(suggestion["priority"])
        self.data_vars["available_at_triage"].set(suggestion["available_at_triage"])
        self.data_vars["status"].set(suggestion["status"])
        self.update_status("Data element suggestion applied")

    def validate_data_inputs(self, for_update=False):
        values = {key: var.get().strip() for key, var in self.data_vars.items()}
        required = ["category", "data_element", "data_type", "source", "priority", "available_at_triage", "status"]
        for field in required:
            if not values[field]:
                messagebox.showwarning("Validation", f"{field.replace('_', ' ').title()} is required.")
                return None
        for index, item in enumerate(self.data_inventory):
            if item["data_element"].strip().lower() == values["data_element"].lower():
                if not for_update or index != self.selected_data_index:
                    messagebox.showwarning("Validation", "Duplicate data element name is not allowed.")
                    return None
        return values

    def add_data_item(self):
        values = self.validate_data_inputs(for_update=False)
        if not values:
            return
        self.data_inventory.append(values)
        self.refresh_data_tree()
        self.clear_data_form()
        self.update_status("Data item added")
        messagebox.showinfo("Added", "Data item added successfully.")

    def update_data_item(self):
        if self.selected_data_index is None:
            messagebox.showwarning("Update", "Select a data item to update.")
            return
        values = self.validate_data_inputs(for_update=True)
        if not values:
            return
        self.data_inventory[self.selected_data_index] = values
        self.refresh_data_tree()
        self.clear_data_form()
        self.update_status("Data item updated")
        messagebox.showinfo("Updated", "Data item updated successfully.")

    def delete_data_item(self):
        if self.selected_data_index is None:
            messagebox.showwarning("Delete", "Select a data item to delete.")
            return
        if messagebox.askyesno("Confirm Delete", "Delete the selected data item?"):
            del self.data_inventory[self.selected_data_index]
            self.refresh_data_tree()
            self.clear_data_form()
            self.update_status("Data item deleted")

    def refresh_data_tree(self):
        for item in self.data_tree.get_children():
            self.data_tree.delete(item)

        search_term = self.data_search_var.get().strip().lower()
        category_filter = self.data_filter_var.get()

        for item in self.data_inventory:
            if category_filter != "All" and item["category"] != category_filter:
                continue
            if search_term and search_term not in item["data_element"].lower():
                continue
            tags = ()
            if item["priority"] == "High" and item["status"] in {"Add", "Exclude"}:
                tags = ("priority_attention",)
            self.data_tree.insert("", "end", values=(
                item["category"], item["data_element"], item["data_type"], item["source"], item["why_it_matters"], item["priority"], item["available_at_triage"], item["status"]
            ), tags=tags)

    def on_data_select(self, event=None):
        selection = self.data_tree.selection()
        if not selection:
            return
        item_id = selection[0]
        values = self.data_tree.item(item_id, "values")
        target_name = values[1]
        for index, item in enumerate(self.data_inventory):
            if item["data_element"] == target_name:
                self.selected_data_index = index
                break
        keys = ["category", "data_element", "data_type", "source", "why_it_matters", "priority", "available_at_triage", "status"]
        for key, value in zip(keys, values):
            self.data_vars[key].set(value)

    def clear_data_form(self):
        for var in self.data_vars.values():
            var.set("")
        self.selected_data_index = None

    def load_default_data_inventory(self):
        if messagebox.askyesno("Load Data Inventory", "Replace current data inventory with full project data?"):
            self.data_inventory = default_data_inventory()
            self.refresh_data_tree()
            self.clear_data_form()
            self.update_status("Default project data inventory loaded")

    def check_data_coverage(self):
        existing = {item["data_element"].strip().lower() for item in self.data_inventory}
        missing = [name for name in REQUIRED_DATA_ELEMENTS if name.strip().lower() not in existing]
        if missing:
            messagebox.showwarning("Data Coverage", "Missing project data elements:\n\n- " + "\n- ".join(missing))
        else:
            messagebox.showinfo("Data Coverage", "All core project data elements are present.")
        self.update_status("Data coverage checked")

    def export_data_csv(self):
        if not self.data_inventory:
            messagebox.showwarning("Export", "No data inventory to export.")
            return
        filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")], title="Export Data Inventory CSV")
        if not filename:
            return
        try:
            with open(filename, "w", newline="", encoding="utf-8") as file:
                writer = csv.DictWriter(file, fieldnames=["category", "data_element", "data_type", "source", "why_it_matters", "priority", "available_at_triage", "status"])
                writer.writeheader()
                writer.writerows(self.data_inventory)
            messagebox.showinfo("Exported", "Data inventory CSV exported successfully.")
            self.update_status("Data inventory CSV exported")
        except Exception as error:
            messagebox.showerror("Export Error", f"Could not export data inventory CSV.\n\n{error}")

    # ------------------------- Patient Case Entry -------------------------
    def build_patient_tab(self):
        self.patient_sections = {}

        patient_banner = self.create_banner(
            self.patient_tab,
            "Patient Review Workspace",
            "Enter the case once, review the AI/CDSS output, optionally apply clinician override, then generate a structured clinical document and keep an audit-ready trail.",
            bg="#12344d",
        )
        patient_banner.pack(fill="x", pady=(0, 8))

        patient_cards_row = ttk.Frame(self.patient_tab)
        patient_cards_row.pack(fill="x", pady=(0, 8))
        self.patient_kpi_vars = {
            "risk": tk.StringVar(value="Not calculated"),
            "pathway": tk.StringVar(value="Not calculated"),
            "confidence": tk.StringVar(value="Not calculated"),
            "missing": tk.StringVar(value="0"),
            "document": tk.StringVar(value="Not generated"),
        }
        self.patient_kpi_value_labels = {}
        patient_card_specs = [
            ("Live Risk", "risk", "Current AI/CDSS risk classification", UI_COLORS["blue_soft"]),
            ("Final Pathway", "pathway", "Current displayed pathway", UI_COLORS["teal_soft"]),
            ("Confidence", "confidence", "Current confidence level", UI_COLORS["green_soft"]),
            ("Missing Critical", "missing", "Number of missing critical fields", UI_COLORS["amber_soft"]),
            ("Document Draft", "document", "Latest generated document status", UI_COLORS["panel"]),
        ]
        for title, key, subtitle, bg in patient_card_specs:
            card, value_label = self.create_metric_card(patient_cards_row, title, self.patient_kpi_vars[key], subtitle=subtitle, bg=bg, width=215)
            card.pack(side="left", fill="x", expand=True, padx=(0, 8))
            self.patient_kpi_value_labels[key] = value_label

        predictive_frame = ttk.LabelFrame(self.patient_tab, text="Predictive Use Case", padding=8)
        predictive_frame.pack(fill="x", pady=(0, 6))
        self.patient_sections["Predictive Use Case"] = predictive_frame

        self.case_target_var = tk.StringVar(value=self.app_data.get("project_target", PREDICTION_TARGETS[1]))
        ttk.Label(predictive_frame, text="Prediction Target:", style="PredictiveLabel.TLabel").grid(row=0, column=0, sticky="w", padx=4, pady=2)
        self.case_target_combo = ttk.Combobox(
            predictive_frame,
            textvariable=self.case_target_var,
            values=PREDICTION_TARGETS,
            state="readonly",
            width=24,
            style="Compact.TCombobox",
        )
        self.case_target_combo.grid(row=0, column=1, sticky="w", padx=4, pady=2)
        self.case_target_combo.bind("<<ComboboxSelected>>", self.on_case_target_change)

        ttk.Label(predictive_frame, text="Prediction Timing:", style="PredictiveLabel.TLabel").grid(row=1, column=0, sticky="nw", padx=4, pady=2)
        ttk.Label(predictive_frame, text=PREDICTION_TIMING_TEXT, style="PredictiveValue.TLabel", wraplength=520, justify="left").grid(row=1, column=1, sticky="w", padx=4, pady=2)

        ttk.Label(predictive_frame, text="Primary Users:", style="PredictiveLabel.TLabel").grid(row=2, column=0, sticky="nw", padx=4, pady=2)
        ttk.Label(predictive_frame, text=PREDICTION_USERS_TEXT, style="PredictiveValue.TLabel", wraplength=520, justify="left").grid(row=2, column=1, sticky="w", padx=4, pady=2)

        ttk.Label(predictive_frame, text="Output Type:", style="PredictiveLabel.TLabel").grid(row=3, column=0, sticky="nw", padx=4, pady=2)
        ttk.Label(predictive_frame, text=PREDICTION_OUTPUT_TEXT, style="PredictiveValue.TLabel", wraplength=520, justify="left").grid(row=3, column=1, sticky="w", padx=4, pady=2)

        ttk.Label(predictive_frame, text="Clinician Oversight:", style="PredictiveLabel.TLabel").grid(row=4, column=0, sticky="nw", padx=4, pady=2)
        ttk.Label(predictive_frame, text=CLINICIAN_OVERSIGHT_TEXT, style="PredictiveValueAccent.TLabel", wraplength=520, justify="left").grid(row=4, column=1, sticky="w", padx=4, pady=2)
        predictive_frame.columnconfigure(1, weight=1)

        form_frame = ttk.LabelFrame(self.patient_tab, text="Patient Case Form", padding=3)
        form_frame.pack(fill="x", pady=(0, 8))
        self.patient_sections["Patient Case Form"] = form_frame

        self.case_vars = {
            "patient_id": tk.StringVar(),
            "patient_name": tk.StringVar(),
            "age": tk.StringVar(),
            "sex": tk.StringVar(),
            "arrival_time": tk.StringVar(value=datetime.now().strftime("%Y-%m-%d %H:%M")),
            "shortness_of_breath": tk.StringVar(),
            "edema": tk.StringVar(),
            "fatigue": tk.StringVar(),
            "chest_pain": tk.StringVar(),
            "systolic_bp": tk.StringVar(),
            "diastolic_bp": tk.StringVar(),
            "heart_rate": tk.StringVar(),
            "respiratory_rate": tk.StringVar(),
            "oxygen_saturation": tk.StringVar(),
            "temperature": tk.StringVar(),
            "previous_admissions": tk.StringVar(),
            "medication_history": tk.StringVar(),
            "comorbidities": tk.StringVar(),
            "severe_dyspnea": tk.StringVar(),
            "confusion": tk.StringVar(),
            "pulmonary_edema_signs": tk.StringVar(),
            "arrhythmia_symptoms": tk.StringVar(),
            "triage_nurse_notes": tk.StringVar(),
            "risk_output": tk.StringVar(value="Medium Risk"),
            "final_pathway": tk.StringVar(value="Final Triage Blocked"),
        }

        yes_no_values = ["No", "Yes"]
        sex_values = ["Male", "Female", "Other"]

        fields = [
            ("Patient ID", "patient_id", None),
            ("Patient Name", "patient_name", None),
            ("Age", "age", None),
            ("Sex", "sex", sex_values),
            ("ED Arrival Time", "arrival_time", None),
            ("Shortness of Breath", "shortness_of_breath", yes_no_values),
            ("Edema", "edema", yes_no_values),
            ("Fatigue", "fatigue", yes_no_values),
            ("Chest Pain", "chest_pain", yes_no_values),
            ("Systolic BP", "systolic_bp", None),
            ("Diastolic BP", "diastolic_bp", None),
            ("Heart Rate", "heart_rate", None),
            ("Respiratory Rate", "respiratory_rate", None),
            ("Oxygen Saturation", "oxygen_saturation", None),
            ("Temperature", "temperature", None),
            ("Previous Admissions", "previous_admissions", None),
            ("Medication History", "medication_history", None),
            ("Comorbidities", "comorbidities", None),
            ("Severe Dyspnea", "severe_dyspnea", yes_no_values),
            ("Confusion", "confusion", yes_no_values),
            ("Pulmonary Edema Signs", "pulmonary_edema_signs", yes_no_values),
            ("Arrhythmia Symptoms", "arrhythmia_symptoms", yes_no_values),
        ]

        compact_labels = {
            "Patient ID": "Patient ID",
            "Patient Name": "Patient Name",
            "Age": "Age",
            "Sex": "Sex",
            "ED Arrival Time": "ED Arrival Time",
            "Shortness of Breath": "Shortness of Breath",
            "Edema": "Edema",
            "Fatigue": "Fatigue",
            "Chest Pain": "Chest Pain",
            "Systolic BP": "Systolic BP",
            "Diastolic BP": "Diastolic BP",
            "Heart Rate": "Heart Rate",
            "Respiratory Rate": "Respiratory Rate",
            "Oxygen Saturation": "Oxygen Saturation",
            "Temperature": "Temperature",
            "Previous Admissions": "Previous Admissions",
            "Medication History": "Medication History",
            "Comorbidities": "Comorbidities",
            "Severe Dyspnea": "Severe Dyspnea",
            "Confusion": "Confusion",
            "Pulmonary Edema Signs": "Pulmonary Edema Signs",
            "Arrhythmia Symptoms": "Arrhythmia Symptoms",
        }

        # Shorten only the labels that usually get visually eaten on smaller windows.
        display_labels = {
            "Shortness of Breath": "Shortness of Breath",
            "Previous Admissions": "Previous Admissions",
            "Pulmonary Edema Signs": "Pulmonary Edema Signs",
            "Medication History": "Medication History",
            "Arrhythmia Symptoms": "Arrhythmia Symptoms",
        }

        compact_labels = {
            "Patient ID": "Patient ID",
            "Patient Name": "Patient Name",
            "Age": "Age",
            "Sex": "Sex",
            "ED Arrival Time": "ED Arrival",
            "Shortness of Breath": "SOB",
            "Edema": "Edema",
            "Fatigue": "Fatigue",
            "Chest Pain": "Chest Pain",
            "Systolic BP": "Sys BP",
            "Diastolic BP": "Dia BP",
            "Heart Rate": "Heart Rate",
            "Respiratory Rate": "Resp Rate",
            "Oxygen Saturation": "O2 Sat",
            "Temperature": "Temp",
            "Previous Admissions": "Prev Adm",
            "Medication History": "Medication",
            "Comorbidities": "Comorbidities",
            "Severe Dyspnea": "Severe Dysp",
            "Confusion": "Confusion",
            "Pulmonary Edema Signs": "Pulm Edema",
            "Arrhythmia Symptoms": "Arrhythmia",
        }

        columns_per_row = 3
        field_row_count = ((len(fields) - 1) // columns_per_row) + 1
        readonly_combo_keys = {"sex", "shortness_of_breath", "edema", "fatigue", "chest_pain", "severe_dyspnea", "confusion", "pulmonary_edema_signs", "arrhythmia_symptoms"}

        for index, (label_text, key, values) in enumerate(fields):
            row = index // columns_per_row
            col = index % columns_per_row
            cell = ttk.Frame(form_frame)
            cell.grid(row=row, column=col, sticky="ew", padx=4, pady=2)
            cell.columnconfigure(1, weight=1)

            shown_label = compact_labels.get(label_text, label_text)
            ttk.Label(
                cell,
                text=shown_label,
                style="PatientField.TLabel",
                width=13,
                anchor="w",
                justify="left"
            ).grid(row=0, column=0, sticky="w", padx=(0, 4), pady=0)

            if values is None:
                wider_fields = {
                    "patient_name": 22,
                    "arrival_time": 22,
                    "medication_history": 22,
                    "comorbidities": 22,
                    "previous_admissions": 18,
                }
                entry_width = wider_fields.get(key, 11)
                ttk.Entry(
                    cell,
                    textvariable=self.case_vars[key],
                    width=entry_width,
                    style="Compact.TEntry"
                ).grid(row=0, column=1, sticky="w")
            else:
                combo_width = 13 if key == "sex" else 12
                combo = ttk.Combobox(
                    cell,
                    textvariable=self.case_vars[key],
                    values=values,
                    width=combo_width,
                    style="Compact.TCombobox"
                )
                combo.grid(row=0, column=1, sticky="w")
                if key in readonly_combo_keys:
                    combo.configure(state="readonly")

        notes_label = ttk.Label(form_frame, text="Triage Nurse Notes", style="PatientField.TLabel")
        notes_label.grid(row=field_row_count, column=0, sticky="w", padx=4, pady=3)
        self.case_notes_text = tk.Text(form_frame, height=3, width=76, wrap="word", font=("Segoe UI", 9))
        self.case_notes_text.grid(row=field_row_count, column=1, columnspan=2, sticky="ew", padx=4, pady=2)
        self.apply_text_panel_style(self.case_notes_text)

        for column in range(columns_per_row):
            form_frame.columnconfigure(column, weight=1, minsize=320)

        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=field_row_count + 1, column=0, columnspan=columns_per_row, sticky="ew", pady=(5, 0))
        ttk.Button(button_frame, text="Generate Risk Classification", style="Compact.TButton", command=self.suggest_risk_output).pack(side="left", padx=4)
        ttk.Button(
    button_frame,
    text="Check Doctor Decision",
    style="Compact.TButton",
    command=self.check_doctor_decision
).pack(side="left", padx=4, pady=4)
        ttk.Button(button_frame, text="Show Risk Criteria", style="Compact.TButton", command=self.show_risk_criteria).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Generate Summary", style="Compact.TButton", command=self.generate_case_summary).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Add Patient Case", style="Compact.TButton", command=self.add_case).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Update Patient Case", style="Compact.TButton", command=self.update_case).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Delete Patient Case", style="Compact.TButton", command=self.delete_case).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Reset Form", style="Compact.TButton", command=self.clear_case_form).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Import Excel", style="Compact.TButton", command=self.import_case_excel).pack(side="left", padx=4)
        ttk.Button(button_frame, text="Save", style="Compact.TButton", command=self.save_all_data).pack(side="right", padx=4)
        ttk.Button(button_frame, text="Export Excel", style="Compact.TButton", command=self.export_case_excel).pack(side="right", padx=4)
        ttk.Button(button_frame, text="Export CSV", style="Compact.TButton", command=self.export_case_csv).pack(side="right", padx=4)

        navigation_frame = ttk.LabelFrame(self.patient_tab, text="Patient Navigation", padding=8)
        navigation_frame.pack(fill="x", pady=(0, 8))
        self.patient_selector_var = tk.StringVar()
        ttk.Label(navigation_frame, text="Choose Patient:").pack(side="left", padx=(0, 6))
        self.patient_selector_combo = ttk.Combobox(navigation_frame, textvariable=self.patient_selector_var, state="readonly", width=22, style="Compact.TCombobox")
        self.patient_selector_combo.pack(side="left", padx=(0, 8))
        self.patient_selector_combo.bind("<<ComboboxSelected>>", self.on_patient_selector_change)
        ttk.Button(navigation_frame, text="Previous Patient", style="Compact.TButton", command=self.load_previous_patient).pack(side="left", padx=4)
        ttk.Button(navigation_frame, text="Next Patient", style="Compact.TButton", command=self.load_next_patient).pack(side="left", padx=4)
        ttk.Label(navigation_frame, text="Tip: choose a patient here instead of scrolling to the bottom table.").pack(side="right", padx=4)

        insights_frame = ttk.LabelFrame(self.patient_tab, text="Decision Support Insights", padding=8)
        insights_frame.pack(fill="x", pady=(0, 8))
        self.patient_sections["Decision Support Insights"] = insights_frame

        self.output_risk_var = tk.StringVar(value="Not calculated")
        self.output_pathway_var = tk.StringVar(value="Not calculated")
        self.risk_score_var = tk.StringVar(value="Not calculated")
        self.early_warning_var = tk.StringVar(value="Not calculated")
        self.recommended_action_var = tk.StringVar(value="Not calculated")
        self.confidence_var = tk.StringVar(value="Not calculated")
        self.human_loop_var = tk.StringVar(value="Human-in-the-Loop: Final decision requires clinician review")
        self.alert_delivery_var = tk.StringVar(value="Not calculated")
        self.alert_fatigue_var = tk.StringVar(value="No alert has been generated yet")
        self.local_acuity_var = tk.StringVar(value="Not calculated")
        self.safety_lock_var = tk.StringVar(value="Safety status not calculated")
        self.triage_gate_var = tk.StringVar(value="No emergency gate evaluated yet")
        self.clinician_confirmation_var = tk.StringVar(value="Pending clinician confirmation")
        self.final_decision_confirmed_var = tk.BooleanVar(value=False)

        ttk.Label(insights_frame, text="Risk Category:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.risk_output_label = ttk.Label(insights_frame, textvariable=self.output_risk_var, style="RiskDefault.TLabel")
        self.risk_output_label.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(insights_frame, text="Risk Score:").grid(row=0, column=2, sticky="w", padx=5, pady=5)
        self.risk_score_label = ttk.Label(insights_frame, textvariable=self.risk_score_var, style="RiskDefault.TLabel")
        self.risk_score_label.grid(row=0, column=3, sticky="w", padx=5, pady=5)

        ttk.Label(insights_frame, text="Early Warning Alert:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.early_warning_label = ttk.Label(insights_frame, textvariable=self.early_warning_var, style="RiskDefault.TLabel", wraplength=280, justify="left")
        self.early_warning_label.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(insights_frame, text="Recommended Action:").grid(row=1, column=2, sticky="w", padx=5, pady=5)
        self.recommended_action_label = ttk.Label(insights_frame, textvariable=self.recommended_action_var, style="RiskDefault.TLabel", wraplength=280, justify="left")
        self.recommended_action_label.grid(row=1, column=3, sticky="w", padx=5, pady=5)

        ttk.Label(insights_frame, text="Final Pathway:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.final_pathway_label = ttk.Label(insights_frame, textvariable=self.output_pathway_var, style="RiskDefault.TLabel")
        self.final_pathway_label.grid(row=2, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(insights_frame, text="Confidence Level:").grid(row=2, column=2, sticky="w", padx=5, pady=5)
        self.confidence_label = ttk.Label(insights_frame, textvariable=self.confidence_var, style="ConfidenceDefault.TLabel")
        self.confidence_label.grid(row=2, column=3, sticky="w", padx=5, pady=5)

        ttk.Label(insights_frame, text="Human-in-the-Loop:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, textvariable=self.human_loop_var, style="HumanLoop.TLabel", wraplength=760, justify="left").grid(row=3, column=1, columnspan=3, sticky="w", padx=5, pady=5)

        ttk.Label(insights_frame, text="Alert Delivery:").grid(row=4, column=0, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, textvariable=self.alert_delivery_var, style="HumanLoop.TLabel", wraplength=320, justify="left").grid(row=4, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, text="Alert Fatigue Status:").grid(row=4, column=2, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, textvariable=self.alert_fatigue_var, wraplength=320, justify="left").grid(row=4, column=3, sticky="w", padx=5, pady=5)

        alert_control_frame = ttk.Frame(insights_frame)
        alert_control_frame.grid(row=5, column=0, columnspan=4, sticky="ew", padx=5, pady=(0, 5))
        self.acknowledge_alert_button = ttk.Button(alert_control_frame, text="Acknowledge Alert", style="Compact.TButton", command=self.acknowledge_current_alert)
        self.acknowledge_alert_button.pack(side="left", padx=(0, 6))
        self.reset_alert_button = ttk.Button(alert_control_frame, text="Reset Suppression", style="Compact.TButton", command=self.reset_alert_suppression)
        self.reset_alert_button.pack(side="left", padx=6)
        ttk.Label(alert_control_frame, text=f"Policy: High Risk = interruptive once per {ALERT_SUPPRESSION_WINDOW_MINUTES} minutes after acknowledgement | Medium Risk = dashboard only | Low Risk = silent", wraplength=760, justify="left").pack(side="left", padx=8)

        ttk.Label(insights_frame, text="Local ED Acuity:").grid(row=6, column=0, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, textvariable=self.local_acuity_var, style="HumanLoop.TLabel", wraplength=320, justify="left").grid(row=6, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, text="Safety Lock:").grid(row=6, column=2, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, textvariable=self.safety_lock_var, style="RiskDefault.TLabel", wraplength=360, justify="left").grid(row=6, column=3, sticky="w", padx=5, pady=5)

        ttk.Label(insights_frame, text="Emergency Gates:").grid(row=7, column=0, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, textvariable=self.triage_gate_var, wraplength=320, justify="left").grid(row=7, column=1, sticky="w", padx=5, pady=5)
        ttk.Label(insights_frame, text="Clinician Confirmation:").grid(row=7, column=2, sticky="w", padx=5, pady=5)
        confirm_frame = ttk.Frame(insights_frame)
        confirm_frame.grid(row=7, column=3, sticky="ew", padx=5, pady=5)
        ttk.Label(confirm_frame, textvariable=self.clinician_confirmation_var, style="HumanLoop.TLabel", wraplength=230, justify="left").pack(side="left", padx=(0, 8))
        ttk.Button(confirm_frame, text="Confirm Final Pathway", style="Compact.TButton", command=self.confirm_final_pathway).pack(side="left")

        ttk.Label(insights_frame, text="Red Flags Detected:").grid(row=8, column=0, sticky="nw", padx=5, pady=5)
        self.red_flags_text = tk.Text(insights_frame, height=3, width=56, wrap="word", font=("Segoe UI", 9))
        self.red_flags_text.grid(row=8, column=1, columnspan=3, sticky="ew", padx=5, pady=5)
        self.apply_text_panel_style(self.red_flags_text)

        ttk.Label(insights_frame, text="Missing Critical Fields:").grid(row=9, column=0, sticky="nw", padx=5, pady=5)
        self.missing_critical_text = tk.Text(insights_frame, height=3, width=56, wrap="word", font=("Segoe UI", 9))
        self.missing_critical_text.grid(row=9, column=1, columnspan=3, sticky="ew", padx=5, pady=5)
        self.apply_text_panel_style(self.missing_critical_text)

        override_frame = ttk.LabelFrame(insights_frame, text="Clinician Override", padding=8)
        override_frame.grid(row=10, column=0, columnspan=4, sticky="ew", padx=5, pady=(8, 5))
        self.override_risk_var = tk.StringVar(value="")
        self.override_pathway_var = tk.StringVar(value="")
        ttk.Label(override_frame, text="Override Risk Output:").grid(row=0, column=0, sticky="w", padx=5, pady=4)
        ttk.Combobox(override_frame, textvariable=self.override_risk_var, values=["", *RISK_CLASSES], state="readonly", width=16, style="Compact.TCombobox").grid(row=0, column=1, sticky="w", padx=5, pady=4)
        ttk.Label(override_frame, text="Override Final Pathway:").grid(row=0, column=2, sticky="w", padx=5, pady=4)
        ttk.Combobox(override_frame, textvariable=self.override_pathway_var, values=["", *FINAL_PATHWAYS], state="readonly", width=16, style="Compact.TCombobox").grid(row=0, column=3, sticky="w", padx=5, pady=4)
        ttk.Label(override_frame, text="Reason for Override:").grid(row=1, column=0, sticky="nw", padx=5, pady=4)
        self.override_reason_text = tk.Text(override_frame, height=3, width=68, wrap="word", font=("Segoe UI", 9))
        self.override_reason_text.grid(row=1, column=1, columnspan=3, sticky="ew", padx=5, pady=4)
        self.apply_text_panel_style(self.override_reason_text)
        ttk.Button(override_frame, text="Apply Override", style="Compact.TButton", command=self.apply_clinician_override).grid(row=0, column=4, rowspan=2, sticky="ns", padx=8, pady=4)
        for c in range(4):
            override_frame.columnconfigure(c, weight=1)

        for column in range(4):
            insights_frame.columnconfigure(column, weight=1)
        self.acknowledge_alert_button.configure(state="disabled")
        self.reset_alert_button.configure(state="disabled")

        summary_frame = ttk.LabelFrame(self.patient_tab, text="Case Summary", padding=8)
        summary_frame.pack(fill="both", pady=(0, 8))
        self.patient_sections["Case Summary"] = summary_frame
        summary_inner = ttk.Frame(summary_frame)
        summary_inner.pack(fill="both", expand=True)
        self.case_summary_text = tk.Text(summary_inner, height=4, wrap="word", font=("Segoe UI", 9))
        self.case_summary_text.pack(side="left", fill="both", expand=True)
        self.apply_text_panel_style(self.case_summary_text)
        summary_scroll = ttk.Scrollbar(summary_inner, orient="vertical", command=self.case_summary_text.yview)
        summary_scroll.pack(side="right", fill="y")
        self.case_summary_text.configure(yscrollcommand=summary_scroll.set)

        evidence_frame = ttk.LabelFrame(self.patient_tab, text="Explainability Panel", padding=8)
        evidence_frame.pack(fill="both", pady=(0, 8), expand=False)
        self.patient_sections["Explainability Panel"] = evidence_frame
        evidence_inner = ttk.Frame(evidence_frame)
        evidence_inner.pack(fill="both", expand=True)
        self.case_evidence_text = tk.Text(evidence_inner, height=8, wrap="word", font=("Segoe UI", 9))
        self.case_evidence_text.pack(side="left", fill="both", expand=True)
        self.apply_text_panel_style(self.case_evidence_text)
        evidence_scroll = ttk.Scrollbar(evidence_inner, orient="vertical", command=self.case_evidence_text.yview)
        evidence_scroll.pack(side="right", fill="y")
        self.case_evidence_text.configure(yscrollcommand=evidence_scroll.set)

        documentation_frame = ttk.LabelFrame(self.patient_tab, text="Clinical Documentation Generator", padding=8)
        documentation_frame.pack(fill="both", pady=(0, 8), expand=False)
        self.patient_sections["Clinical Documentation"] = documentation_frame

        self.user_role_var = tk.StringVar(value="Emergency Physician")
        self.document_type_var = tk.StringVar(value=DOCUMENT_TYPES[0])
        self.doc_accuracy_var = tk.BooleanVar(value=False)
        self.doc_completeness_var = tk.BooleanVar(value=False)
        self.doc_traceability_var = tk.BooleanVar(value=False)
        self.doc_safety_var = tk.BooleanVar(value=False)

        controls_row = ttk.Frame(documentation_frame)
        controls_row.pack(fill="x", pady=(0, 6))
        ttk.Label(controls_row, text="Reviewing Role:").pack(side="left", padx=(0, 5))
        ttk.Combobox(controls_row, textvariable=self.user_role_var, values=CLINICIAN_ROLES, state="readonly", width=20, style="Compact.TCombobox").pack(side="left", padx=(0, 10))
        ttk.Label(controls_row, text="Document Type:").pack(side="left", padx=(0, 5))
        ttk.Combobox(controls_row, textvariable=self.document_type_var, values=DOCUMENT_TYPES, state="readonly", width=18, style="Compact.TCombobox").pack(side="left", padx=(0, 10))
        ttk.Button(controls_row, text="Generate Document", style="Compact.TButton", command=self.generate_clinical_document).pack(side="left", padx=4)
        ttk.Button(controls_row, text="Copy", style="Compact.TButton", command=self.copy_generated_document).pack(side="left", padx=4)
        ttk.Button(controls_row, text="Save TXT", style="Compact.TButton", command=self.save_generated_document_txt).pack(side="left", padx=4)

        checklist_frame = ttk.Frame(documentation_frame)
        checklist_frame.pack(fill="x", pady=(0, 6))
        ttk.Checkbutton(checklist_frame, text="Accuracy reviewed", variable=self.doc_accuracy_var).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(checklist_frame, text="Completeness checked", variable=self.doc_completeness_var).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(checklist_frame, text="Traceability included", variable=self.doc_traceability_var).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(checklist_frame, text="Safety statement included", variable=self.doc_safety_var).pack(side="left", padx=(0, 12))

        doc_inner = ttk.Frame(documentation_frame)
        doc_inner.pack(fill="both", expand=True)
        self.documentation_text = tk.Text(doc_inner, height=14, wrap="word", font=("Segoe UI", 9))
        self.documentation_text.pack(side="left", fill="both", expand=True)
        self.apply_text_panel_style(self.documentation_text)
        doc_scroll = ttk.Scrollbar(doc_inner, orient="vertical", command=self.documentation_text.yview)
        doc_scroll.pack(side="right", fill="y")
        self.documentation_text.configure(yscrollcommand=doc_scroll.set)

        search_frame = ttk.Frame(self.patient_tab)
        search_frame.pack(fill="x", pady=(0, 8))
        ttk.Label(search_frame, text="Search Patient:").pack(side="left", padx=(0, 5))
        self.case_search_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.case_search_var, width=28, style="Compact.TEntry").pack(side="left")
        self.case_search_var.trace_add("write", lambda *args: self.refresh_case_tree())

        tree_frame = ttk.LabelFrame(self.patient_tab, text="Saved Patient Cases", padding=8)
        tree_frame.pack(fill="both", expand=True)
        self.patient_sections["Saved Patient Cases"] = tree_frame
        self._bind_scroll_to_children(self.patient_tab)

        case_columns = (
            "patient_id", "patient_name", "age", "sex", "arrival_time",
            "risk_output", "final_pathway", "target_output"
        )
        self.case_tree = ttk.Treeview(tree_frame, columns=case_columns, show="headings")
        headings = {
            "patient_id": "Patient ID",
            "patient_name": "Patient Name",
            "age": "Age",
            "sex": "Sex",
            "arrival_time": "Arrival Time",
            "risk_output": "Risk Output",
            "final_pathway": "Final Pathway",
            "target_output": "Prediction Target",
        }
        for column in case_columns:
            self.case_tree.heading(column, text=headings[column])
        self.case_tree.column("patient_id", width=120)
        self.case_tree.column("patient_name", width=160)
        self.case_tree.column("age", width=60, anchor="center")
        self.case_tree.column("sex", width=80, anchor="center")
        self.case_tree.column("arrival_time", width=150)
        self.case_tree.column("risk_output", width=96, anchor="center")
        self.case_tree.column("final_pathway", width=120, anchor="center")
        self.case_tree.column("target_output", width=250)
        self.case_tree.pack(side="left", fill="both", expand=True)
        self.case_tree.bind("<<TreeviewSelect>>", self.on_case_select)
        self.case_tree.tag_configure("high_risk", background="#ffe1e1")
        self.case_tree.tag_configure("medium_risk", background="#fff3cf")
        self.case_tree.tag_configure("low_risk", background="#e7f9ee")

        case_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.case_tree.yview)
        case_scroll.pack(side="right", fill="y")
        self.case_tree.configure(yscrollcommand=case_scroll.set)
        self.update_patient_dashboard_cards({"missing_critical": []})

    def update_document_checklist(self, evidence, document_text):
        self.doc_accuracy_var.set(False)
        self.doc_completeness_var.set(len(evidence.get("missing_critical", [])) == 0)
        self.doc_traceability_var.set(True)
        self.doc_safety_var.set("NEEDS CLINICIAN REVIEW" in document_text)

    def generate_clinical_document(self):
        patient_id = self.case_vars["patient_id"].get().strip()
        patient_name = self.case_vars["patient_name"].get().strip()
        if not patient_id or not patient_name:
            messagebox.showwarning("Documentation", "Enter at least Patient ID and Patient Name first.")
            return

        evidence = self.build_risk_evidence()
        display_risk = self.override_risk_var.get().strip() or evidence["risk_output"]
        display_pathway = self.override_pathway_var.get().strip() or evidence["final_pathway"]
        document_type = self.document_type_var.get().strip() or DOCUMENT_TYPES[0]
        role = self.user_role_var.get().strip() or "Clinician"
        triage_notes = self.case_notes_text.get("1.0", "end").strip()
        trigger_summary = "; ".join(evidence["triggers"]) if evidence["triggers"] else "No major triggers recorded"
        vitals_line = (
            f"BP {self.case_vars['systolic_bp'].get()}/{self.case_vars['diastolic_bp'].get()} mmHg, "
            f"HR {self.case_vars['heart_rate'].get()} bpm, RR {self.case_vars['respiratory_rate'].get()}/min, "
            f"SpO2 {self.case_vars['oxygen_saturation'].get()}%, Temp {self.case_vars['temperature'].get()}°C"
        )
        symptom_line = (
            f"Shortness of breath={self.case_vars['shortness_of_breath'].get()}, "
            f"Edema={self.case_vars['edema'].get()}, Fatigue={self.case_vars['fatigue'].get()}, "
            f"Chest pain={self.case_vars['chest_pain'].get()}"
        )
        red_flags = evidence.get("red_flags_detected", [])
        red_flags_line = ", ".join(red_flags) if red_flags else "No immediate red flags detected"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

        common_header = [
            f"{document_type.upper()} DRAFT",
            f"Generated by: {role}",
            f"Generated on: {timestamp}",
            f"Patient ID: {patient_id}",
            f"Patient Name: {patient_name}",
            f"Prediction Target: {self.get_active_prediction_target()}",
            "",
        ]

        if document_type == "SOAP Note":
            body_lines = [
                "S:",
                f"Patient reported: {symptom_line}.",
                f"Triage nurse notes: {triage_notes or 'No additional free-text notes documented.'}",
                "",
                "O:",
                f"Vitals: {vitals_line}.",
                f"Comorbidities: {self.case_vars['comorbidities'].get() or 'Not documented'}.",
                f"Medication history: {self.case_vars['medication_history'].get() or 'Not documented'}.",
                f"Red flags: {red_flags_line}.",
                "",
                "A:",
                f"AI-assisted triage output: {display_risk} with score {evidence['risk_score_text']} and confidence {evidence['confidence_level']}.",
                f"Key triggers: {trigger_summary}.",
                "",
                "P:",
                f"Recommended action: {evidence['recommended_action']}.",
                f"Final pathway for review: {display_pathway}.",
            ]
        elif document_type == "Discharge Summary":
            body_lines = [
                "Encounter Summary:",
                f"Heart-failure triage review completed at {ENTRY_POINT}.",
                f"Presenting symptoms: {symptom_line}.",
                f"Key clinical findings: {vitals_line}.",
                "",
                "AI/CDSS Summary:",
                f"Risk output: {display_risk}.",
                f"Suggested pathway: {display_pathway}.",
                f"Recommended next action: {evidence['recommended_action']}.",
                f"Reason summary: {trigger_summary}.",
                "",
                "Follow-up / Safety:",
                "Review medication adherence, verify follow-up timing, and provide return precautions for worsening dyspnea, chest pain, hypoxia, or confusion.",
            ]
        else:
            body_lines = [
                "Reason for Referral:",
                f"Requesting specialist review after AI-assisted triage output of {display_risk}.",
                "",
                "Key Findings:",
                f"Symptoms: {symptom_line}.",
                f"Vitals: {vitals_line}.",
                f"Red flags: {red_flags_line}.",
                "",
                "Requested Review:",
                f"Please review for pathway '{display_pathway}' and action '{evidence['recommended_action']}'.",
                f"AI trigger summary: {trigger_summary}.",
            ]

        footer_lines = [
            "",
            "Traceability:",
            "Source data used: triage form fields, vital signs, medication history, comorbidities, red-flag indicators, and triage nurse notes.",
            "",
            "SAFETY NOTICE: NEEDS CLINICIAN REVIEW BEFORE USE.",
        ]

        document_text = "\n".join(common_header + body_lines + footer_lines)
        self.documentation_text.delete("1.0", "end")
        self.documentation_text.insert("1.0", document_text)
        self.update_document_checklist(evidence, document_text)
        self.update_patient_dashboard_cards(evidence)

        if self.selected_case_index is not None and 0 <= self.selected_case_index < len(self.patient_cases):
            self.patient_cases[self.selected_case_index]["reviewer_role"] = role
            self.patient_cases[self.selected_case_index]["generated_document_type"] = document_type
            self.patient_cases[self.selected_case_index]["generated_document_text"] = document_text
            self.patient_cases[self.selected_case_index]["doc_accuracy_checked"] = "Yes" if self.doc_accuracy_var.get() else "No"
            self.patient_cases[self.selected_case_index]["doc_completeness_checked"] = "Yes" if self.doc_completeness_var.get() else "No"
            self.patient_cases[self.selected_case_index]["doc_traceability_checked"] = "Yes" if self.doc_traceability_var.get() else "No"
            self.patient_cases[self.selected_case_index]["doc_safety_checked"] = "Yes" if self.doc_safety_var.get() else "No"

        self.append_audit_log(
            event_type="DOCUMENT_GENERATED",
            clinician_action="generated",
            notes=f"{document_type} draft generated for clinician review",
            document_type=document_type,
        )
        self.update_status(f"{document_type} generated")

    def copy_generated_document(self):
        document_text = self.documentation_text.get("1.0", "end").strip()
        if not document_text:
            messagebox.showwarning("Copy", "No generated document to copy.")
            return
        self.clipboard_clear()
        self.clipboard_append(document_text)
        self.update_status("Generated document copied to clipboard")

    def save_generated_document_txt(self):
        document_text = self.documentation_text.get("1.0", "end").strip()
        if not document_text:
            messagebox.showwarning("Save", "No generated document to save.")
            return
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt")],
            title="Save Generated Clinical Document",
        )
        if not filename:
            return
        try:
            with open(filename, "w", encoding="utf-8") as file:
                file.write(document_text)
            self.update_status("Generated document saved to TXT")
        except Exception as error:
            messagebox.showerror("Save Error", f"Could not save the generated document.\n\n{error}")

    def show_risk_criteria(self):
        target = self.get_active_prediction_target()
        lines = [
            "Presentation-aligned risk approach:",
            "- Risk score quantifies deterioration likelihood on a 0-100 scale",
            "- Risk category classifies the patient as Low / Medium / High",
            "- Early warning alert highlights urgency for the clinician",
            "- Recommended clinical action summarizes the next step",
            "",
            f"Current prediction target: {target}",
            "",
            "High Risk classification:",
            "- Typical score band: 70-100",
            "- Typical alert: Immediate early warning alert for urgent intervention",
            "- Typical action: urgent escalation, senior clinician review, and likely urgent care or admission",
            "- Main triggers: severe dyspnea, confusion, pulmonary edema signs, arrhythmia symptoms, chest pain, SBP < 90, SpO2 < 90%, RR >= 30, HR >= 130",
            "",
            "Medium Risk classification:",
            "- Typical score band: 40-69",
            "- Typical alert: Urgent review needed",
            "- Typical action: prompt physician assessment, close monitoring, and escalation if the patient worsens",
            "- Main triggers: borderline vital sign abnormalities, ongoing symptom burden, or history factors without clear severe instability",
            "",
            "Low Risk classification:",
            "- Typical score band: 0-39",
            "- Typical alert: No immediate early warning alert",
            "- Typical action: routine clinician evaluation with standard monitoring and follow-up planning",
            "- Main triggers: no major red flags and overall stable entered data",
            "",
            "Human-in-the-Loop:",
            "- The AI suggests a risk category, score, alert, and action",
            "- The clinician validates the result and can override it before confirming the final pathway",
        ]
        messagebox.showinfo("Presentation-Aligned Risk Approach", "\n".join(lines))

    def to_int_or_none(self, value):
        text_value = str(value).strip()
        if text_value == "":
            return None
        try:
            return int(float(text_value))
        except ValueError:
            return None

    def to_float_or_none(self, value):
        text_value = str(value).strip()
        if text_value == "":
            return None
        try:
            return float(text_value)
        except ValueError:
            return None

    def set_confidence_style(self, confidence_level):
        style_map = {
            "High Confidence": "ConfidenceHigh.TLabel",
            "Moderate Confidence": "ConfidenceModerate.TLabel",
            "Low Confidence": "ConfidenceLow.TLabel",
        }
        self.confidence_label.configure(style=style_map.get(confidence_level, "ConfidenceDefault.TLabel"))

    def set_risk_style(self, risk_output):
        style_map = {
            "High Risk": "RiskHigh.TLabel",
            "Medium Risk": "RiskMedium.TLabel",
            "Low Risk": "RiskLow.TLabel",
        }
        style_name = style_map.get(risk_output, "RiskDefault.TLabel")
        self.risk_output_label.configure(style=style_name)
        self.final_pathway_label.configure(style=style_name)
        if hasattr(self, "risk_score_label"):
            self.risk_score_label.configure(style=style_name)
        if hasattr(self, "early_warning_label"):
            self.early_warning_label.configure(style=style_name)
        if hasattr(self, "recommended_action_label"):
            self.recommended_action_label.configure(style=style_name)

    def confirm_final_pathway(self):
        """Mandatory human confirmation before any saved final pathway."""
        evidence = self.build_risk_evidence()
        display_pathway = self.override_pathway_var.get().strip() or evidence.get("final_pathway", "")
        display_risk = self.override_risk_var.get().strip() or evidence.get("risk_output", "")

        if evidence.get("invalid_ranges"):
            messagebox.showerror(
                "Clinician Confirmation Blocked",
                "Final pathway cannot be confirmed because invalid vital-sign ranges were detected.\n\n"
                + "\n".join(evidence.get("invalid_ranges", [])),
            )
            return

        if evidence.get("missing_critical") and not evidence.get("emergency_gate_active"):
            messagebox.showwarning(
                "Clinician Confirmation Blocked",
                "Final triage is blocked until critical data are completed.\n\nMissing:\n- "
                + "\n- ".join(evidence.get("missing_critical", [])),
            )
            return

        discharge_locked, reasons = discharge_is_locked(evidence, clinician_confirmed=True)
        if "discharge" in display_pathway.lower() and discharge_locked:
            messagebox.showerror(
                "Do Not Discharge Lock",
                "Discharge cannot be confirmed while a safety lock is active.\n\nReasons:\n- " + "\n- ".join(reasons),
            )
            return

        self.final_decision_confirmed_var.set(True)
        self.clinician_confirmation_var.set("Confirmed by clinician")
        evidence = self.build_risk_evidence()
        self.update_decision_support_panels(evidence)
        self.append_audit_log(
            event_type="FINAL_PATHWAY_CONFIRMED",
            clinician_action="confirmed",
            notes=f"Clinician confirmed pathway: {display_pathway} | risk: {display_risk}",
        )
        self.update_status("Final pathway confirmed by clinician")
        messagebox.showinfo("Confirmed", "Final pathway confirmed by clinician. Saving is now allowed if other validation checks pass.")

    def apply_clinician_override(self):
        override_risk = self.override_risk_var.get().strip()
        override_pathway = self.override_pathway_var.get().strip()
        override_reason = self.override_reason_text.get("1.0", "end").strip()
        if not override_risk and not override_pathway:
            messagebox.showwarning("Clinician Override", "Select an override value first.")
            return
        if not override_reason:
            messagebox.showwarning("Clinician Override", "Please enter a reason for override.")
            return
        if override_risk:
            self.case_vars["risk_output"].set(override_risk)
            self.output_risk_var.set(override_risk)
            self.set_risk_style(override_risk)
        if override_pathway:
            self.case_vars["final_pathway"].set(override_pathway)
            self.output_pathway_var.set(override_pathway)
        if hasattr(self, "final_decision_confirmed_var"):
            self.final_decision_confirmed_var.set(False)
        if hasattr(self, "clinician_confirmation_var"):
            self.clinician_confirmation_var.set("Pending clinician confirmation after override")
        self.human_loop_var.set("Human-in-the-Loop: clinician override applied; final confirmation is required before saving")
        current_text = self.case_evidence_text.get("1.0", "end").strip()
        if current_text:
            current_text += "\n\nClinician override applied:\n"
        else:
            current_text = "Clinician override applied:\n"
        if override_risk:
            current_text += f"- Override Risk Output: {override_risk}\n"
        if override_pathway:
            current_text += f"- Override Final Pathway: {override_pathway}\n"
        current_text += f"- Reason: {override_reason}"
        self.case_evidence_text.delete("1.0", "end")
        self.case_evidence_text.insert("1.0", current_text)
        if self.selected_case_index is not None and 0 <= self.selected_case_index < len(self.patient_cases):
            self.patient_cases[self.selected_case_index]["override_risk_output"] = override_risk
            self.patient_cases[self.selected_case_index]["override_final_pathway"] = override_pathway
            self.patient_cases[self.selected_case_index]["override_reason"] = override_reason
            self.patient_cases[self.selected_case_index]["risk_output"] = self.output_risk_var.get()
            self.patient_cases[self.selected_case_index]["final_pathway"] = self.output_pathway_var.get()
            self.refresh_case_tree()
        self.append_audit_log(
            event_type="CLINICIAN_OVERRIDE_APPLIED",
            clinician_action="overrode",
            notes=override_reason,
        )
        self.update_status("Clinician override applied")
    def update_decision_support_panels(self, evidence):
        # Base calculated outputs
        calculated_risk = evidence.get("risk_output", "Not calculated")
        calculated_pathway = evidence.get("final_pathway", "Not calculated")

        # Apply clinician override for display only when present for the current case
        override_risk = self.override_risk_var.get().strip() if hasattr(self, "override_risk_var") else ""
        override_pathway = self.override_pathway_var.get().strip() if hasattr(self, "override_pathway_var") else ""
        override_reason = self.override_reason_text.get("1.0", "end").strip() if hasattr(self, "override_reason_text") else ""

        display_risk = override_risk or calculated_risk
        display_pathway = override_pathway or calculated_pathway

        self.output_risk_var.set(display_risk)
        self.output_pathway_var.set(display_pathway)
        self.risk_score_var.set(evidence.get("risk_score_text", "Not calculated"))
        self.early_warning_var.set(evidence.get("early_warning_alert", "Not calculated"))
        self.recommended_action_var.set(evidence.get("recommended_action", "Not calculated"))
        self.set_risk_style(display_risk)

        if override_risk or override_pathway:
            self.human_loop_var.set("Human-in-the-Loop: clinician override applied; final confirmation is required before saving")
        else:
            self.human_loop_var.set("Human-in-the-Loop: AI suggests the result, but clinician review is required before the final pathway")

        self.confidence_var.set(evidence.get("confidence_level", "Not calculated"))
        self.set_confidence_style(evidence.get("confidence_level", ""))
        if hasattr(self, "alert_delivery_var"):
            self.alert_delivery_var.set(evidence.get("alert_delivery_mode", "Not calculated"))
        if hasattr(self, "alert_fatigue_var"):
            self.alert_fatigue_var.set(evidence.get("alert_fatigue_status", "No alert fatigue status available"))
        if hasattr(self, "local_acuity_var"):
            self.local_acuity_var.set(evidence.get("local_ed_acuity", "Not calculated"))
        if hasattr(self, "safety_lock_var"):
            self.safety_lock_var.set(evidence.get("safety_lock_status", "Safety status not calculated"))
        if hasattr(self, "triage_gate_var"):
            gates = evidence.get("emergency_gates", [])
            self.triage_gate_var.set("Active: " + "; ".join(gates[:3]) if gates else "No emergency gate active")
        if hasattr(self, "clinician_confirmation_var"):
            self.clinician_confirmation_var.set("Confirmed by clinician" if self.final_decision_confirmed_var.get() else "Pending clinician confirmation")
        if hasattr(self, "acknowledge_alert_button"):
            if evidence.get("suppression_active"):
                self.acknowledge_alert_button.configure(state="disabled")
            elif display_risk in ("High Risk", "Medium Risk"):
                self.acknowledge_alert_button.configure(state="normal")
            else:
                self.acknowledge_alert_button.configure(state="disabled")
        if hasattr(self, "reset_alert_button"):
            self.reset_alert_button.configure(state="normal" if self.get_current_patient_key() in getattr(self, "alert_state_by_patient", {}) else "disabled")

        # Red flags panel
        if hasattr(self, "red_flags_text"):
            self.red_flags_text.delete("1.0", "end")
            red_flags = evidence.get("red_flags_detected", [])
            if red_flags:
                for item in red_flags:
                    self.red_flags_text.insert("end", f"- {item}\n")
            else:
                self.red_flags_text.insert("1.0", "No red flags detected.")

        self.missing_critical_text.delete("1.0", "end")
        critical_missing = evidence.get("missing_critical", [])
        invalid_ranges = evidence.get("invalid_ranges", [])
        if critical_missing or invalid_ranges:
            for item in critical_missing:
                self.missing_critical_text.insert("end", f"- Missing: {item}\n")
            for item in invalid_ranges:
                self.missing_critical_text.insert("end", f"- Invalid: {item}\n")
        else:
            self.missing_critical_text.insert("1.0", "No missing critical fields or invalid vital ranges detected.")

        evidence_text = evidence.get("evidence_text", "")
        if override_reason and (override_risk or override_pathway):
            evidence_text = evidence_text.rstrip() + "\n\nClinician override applied:\n"
            if override_risk:
                evidence_text += f"- Override Risk Output: {override_risk}\n"
            if override_pathway:
                evidence_text += f"- Override Final Pathway: {override_pathway}\n"
            evidence_text += f"- Reason: {override_reason}"

        self.case_evidence_text.delete("1.0", "end")
        self.case_evidence_text.insert("1.0", evidence_text)
        self.update_patient_dashboard_cards(evidence)

    def build_risk_evidence(self):
        systolic_bp = self.to_int_or_none(self.case_vars["systolic_bp"].get())
        diastolic_bp = self.to_int_or_none(self.case_vars["diastolic_bp"].get())
        heart_rate = self.to_int_or_none(self.case_vars["heart_rate"].get())
        respiratory_rate = self.to_int_or_none(self.case_vars["respiratory_rate"].get())
        oxygen_saturation = self.to_int_or_none(self.case_vars["oxygen_saturation"].get())
        temperature = self.to_float_or_none(self.case_vars["temperature"].get())
        age = self.to_int_or_none(self.case_vars["age"].get())
        triage_notes = self.case_notes_text.get("1.0", "end").strip()
        previous_admissions = None
        try:
            previous_admissions = self.parse_previous_admissions(self.case_vars["previous_admissions"].get())
        except Exception:
            previous_admissions = None

        safety_snapshot = {
            "age": self.case_vars["age"].get(),
            "systolic_bp": self.case_vars["systolic_bp"].get(),
            "diastolic_bp": self.case_vars["diastolic_bp"].get(),
            "heart_rate": self.case_vars["heart_rate"].get(),
            "respiratory_rate": self.case_vars["respiratory_rate"].get(),
            "oxygen_saturation": self.case_vars["oxygen_saturation"].get(),
            "temperature": self.case_vars["temperature"].get(),
            "severe_dyspnea": self.case_vars["severe_dyspnea"].get(),
            "confusion": self.case_vars["confusion"].get(),
            "pulmonary_edema_signs": self.case_vars["pulmonary_edema_signs"].get(),
            "arrhythmia_symptoms": self.case_vars["arrhythmia_symptoms"].get(),
            "chest_pain": self.case_vars["chest_pain"].get(),
        }
        invalid_ranges = validate_vital_ranges_snapshot(safety_snapshot)
        emergency_gates = evaluate_emergency_gates_snapshot(safety_snapshot)
        emergency_gate_active = bool(emergency_gates)

        missing = []
        if age is None:
            missing.append("Age")
        if systolic_bp is None:
            missing.append("Systolic BP")
        if diastolic_bp is None:
            missing.append("Diastolic BP")
        if heart_rate is None:
            missing.append("Heart rate")
        if respiratory_rate is None:
            missing.append("Respiratory rate")
        if oxygen_saturation is None:
            missing.append("Oxygen saturation")
        if temperature is None:
            missing.append("Temperature")
        if previous_admissions is None:
            missing.append("Previous admissions")

        critical_missing = []
        critical_map = [
            (age is None, "Age"),
            (systolic_bp is None, "Systolic BP"),
            (diastolic_bp is None, "Diastolic BP"),
            (heart_rate is None, "Heart rate"),
            (respiratory_rate is None, "Respiratory rate"),
            (oxygen_saturation is None, "Oxygen saturation"),
            (temperature is None, "Temperature"),
        ]
        for is_missing, label in critical_map:
            if is_missing:
                critical_missing.append(label)

        symptoms_yes = []
        history_yes = []
        red_flags_yes = []
        explainability_symptoms = []
        explainability_vitals = []
        explainability_history = []
        explainability_red_flags = []

        severe_dyspnea = self.case_vars["severe_dyspnea"].get() == "Yes"
        confusion = self.case_vars["confusion"].get() == "Yes"
        pulmonary_edema = self.case_vars["pulmonary_edema_signs"].get() == "Yes"
        arrhythmia = self.case_vars["arrhythmia_symptoms"].get() == "Yes"
        chest_pain = self.case_vars["chest_pain"].get() == "Yes"
        sob = self.case_vars["shortness_of_breath"].get() == "Yes"
        edema = self.case_vars["edema"].get() == "Yes"
        fatigue = self.case_vars["fatigue"].get() == "Yes"
        comorbidities_text = self.case_vars["comorbidities"].get().strip()
        medication_history_text = self.case_vars["medication_history"].get().strip()

        if sob:
            symptoms_yes.append("Shortness of breath = Yes")
            explainability_symptoms.append("Shortness of breath is present")
        if edema:
            symptoms_yes.append("Edema = Yes")
            explainability_symptoms.append("Edema is present")
        if fatigue:
            symptoms_yes.append("Fatigue = Yes")
            explainability_symptoms.append("Fatigue is present")
        if chest_pain:
            symptoms_yes.append("Chest pain = Yes")
            explainability_symptoms.append("Chest pain is present")

        if severe_dyspnea:
            red_flags_yes.append("Severe dyspnea = Yes")
            explainability_red_flags.append("Severe dyspnea was detected")
        if confusion:
            red_flags_yes.append("Confusion = Yes")
            explainability_red_flags.append("Confusion was detected")
        if pulmonary_edema:
            red_flags_yes.append("Pulmonary edema signs = Yes")
            explainability_red_flags.append("Pulmonary edema signs were detected")
        if arrhythmia:
            red_flags_yes.append("Arrhythmia symptoms = Yes")
            explainability_red_flags.append("Arrhythmia symptoms were detected")

        if previous_admissions is not None and previous_admissions >= 1:
            history_yes.append(f"Previous admissions = {previous_admissions}")
            explainability_history.append(f"Previous admissions count is {previous_admissions}")
        if comorbidities_text:
            history_yes.append(f"Comorbidities recorded = {comorbidities_text}")
            explainability_history.append(f"Comorbidities recorded: {comorbidities_text}")
        if medication_history_text:
            history_yes.append("Medication history recorded")
            explainability_history.append("Medication history is available")

        target = self.get_active_prediction_target()
        high_reasons = []
        medium_reasons = []
        low_reasons = []
        criteria_used = []

        # High-risk emergency criteria
        if severe_dyspnea:
            high_reasons.append("Severe dyspnea suggests marked respiratory distress")
            criteria_used.append("Severe dyspnea = Yes")
        if confusion:
            high_reasons.append("Confusion may indicate severe physiological compromise")
            criteria_used.append("Confusion = Yes")
        if pulmonary_edema:
            high_reasons.append("Pulmonary edema signs suggest acute decompensation")
            criteria_used.append("Pulmonary edema signs = Yes")
        if arrhythmia:
            high_reasons.append("Arrhythmia symptoms may indicate unstable cardiac status")
            criteria_used.append("Arrhythmia symptoms = Yes")
        if chest_pain:
            high_reasons.append("Chest pain increases concern for urgent emergency review")
            criteria_used.append("Chest pain = Yes")
        if systolic_bp is not None and systolic_bp < 90:
            high_reasons.append(f"Systolic BP {systolic_bp} is below 90")
            criteria_used.append("Systolic BP < 90")
            explainability_vitals.append(f"Systolic BP is low at {systolic_bp}")
        if oxygen_saturation is not None and oxygen_saturation < 90:
            high_reasons.append(f"Oxygen saturation {oxygen_saturation}% is below 90%")
            criteria_used.append("Oxygen saturation < 90%")
            explainability_vitals.append(f"Oxygen saturation is low at {oxygen_saturation}%")
        if respiratory_rate is not None and respiratory_rate >= 30:
            high_reasons.append(f"Respiratory rate {respiratory_rate} is 30 or higher")
            criteria_used.append("Respiratory rate >= 30")
            explainability_vitals.append(f"Respiratory rate is high at {respiratory_rate}")
        if heart_rate is not None and heart_rate >= 130:
            high_reasons.append(f"Heart rate {heart_rate} is severely elevated")
            criteria_used.append("Heart rate >= 130")
            explainability_vitals.append(f"Heart rate is high at {heart_rate}")

        if target == "Risk of same-day clinical deterioration":
            if systolic_bp is not None and 90 <= systolic_bp < 100:
                medium_reasons.append(f"Borderline systolic BP {systolic_bp} may precede same-day deterioration")
                criteria_used.append("Systolic BP 90-99")
                explainability_vitals.append(f"Borderline systolic BP at {systolic_bp}")
            if oxygen_saturation is not None and 90 <= oxygen_saturation < 94:
                medium_reasons.append(f"Oxygen saturation {oxygen_saturation}% is below 94%")
                criteria_used.append("Oxygen saturation 90-93%")
                explainability_vitals.append(f"Oxygen saturation at {oxygen_saturation}%")
            if heart_rate is not None and 110 < heart_rate < 130:
                medium_reasons.append(f"Heart rate {heart_rate} is above 110")
                criteria_used.append("Heart rate 111-129")
                explainability_vitals.append(f"Heart rate at {heart_rate}")
            if respiratory_rate is not None and 22 < respiratory_rate < 30:
                medium_reasons.append(f"Respiratory rate {respiratory_rate} is above 22")
                criteria_used.append("Respiratory rate 23-29")
                explainability_vitals.append(f"Respiratory rate at {respiratory_rate}")
            if sob:
                medium_reasons.append("Shortness of breath is present")
                criteria_used.append("Shortness of breath = Yes")
            if edema:
                medium_reasons.append("Edema is present")
                criteria_used.append("Edema = Yes")
            if fatigue:
                medium_reasons.append("Fatigue supports current clinical burden")
                criteria_used.append("Fatigue = Yes")

        elif target == "Risk of urgent escalation during the same visit":
            if oxygen_saturation is not None and 90 <= oxygen_saturation < 94:
                medium_reasons.append(f"Oxygen saturation {oxygen_saturation}% may require escalation during the visit")
                criteria_used.append("Oxygen saturation 90-93%")
                explainability_vitals.append(f"Oxygen saturation at {oxygen_saturation}%")
            if respiratory_rate is not None and 24 < respiratory_rate < 30:
                medium_reasons.append(f"Respiratory rate {respiratory_rate} is above 24")
                criteria_used.append("Respiratory rate 25-29")
                explainability_vitals.append(f"Respiratory rate at {respiratory_rate}")
            if heart_rate is not None and 115 < heart_rate < 130:
                medium_reasons.append(f"Heart rate {heart_rate} is above 115")
                criteria_used.append("Heart rate 116-129")
                explainability_vitals.append(f"Heart rate at {heart_rate}")
            if systolic_bp is not None and 90 <= systolic_bp < 100:
                medium_reasons.append(f"Systolic BP {systolic_bp} is below 100")
                criteria_used.append("Systolic BP 90-99")
                explainability_vitals.append(f"Borderline systolic BP at {systolic_bp}")
            if sob:
                medium_reasons.append("Shortness of breath increases urgency in the emergency setting")
                criteria_used.append("Shortness of breath = Yes")
            if edema:
                medium_reasons.append("Edema supports heart failure congestion")
                criteria_used.append("Edema = Yes")

        else:
            if previous_admissions is not None and previous_admissions >= 2:
                medium_reasons.append(f"Previous admissions count {previous_admissions} suggests readmission risk")
                criteria_used.append("Previous admissions >= 2")
            elif previous_admissions is not None and previous_admissions == 1:
                medium_reasons.append("At least one previous admission increases readmission concern")
                criteria_used.append("Previous admissions = 1")
            if age is not None and age >= 70:
                medium_reasons.append(f"Age {age} adds vulnerability to near-term hospitalization")
                criteria_used.append("Age >= 70")
            if sob:
                medium_reasons.append("Shortness of breath suggests ongoing symptom burden")
                criteria_used.append("Shortness of breath = Yes")
            if edema:
                medium_reasons.append("Edema suggests persistent congestion")
                criteria_used.append("Edema = Yes")
            if fatigue:
                medium_reasons.append("Fatigue suggests limited functional reserve")
                criteria_used.append("Fatigue = Yes")
            if comorbidities_text:
                medium_reasons.append("Comorbidities are present")
                criteria_used.append("Comorbidities recorded")
            if oxygen_saturation is not None and 90 <= oxygen_saturation < 94:
                medium_reasons.append(f"Oxygen saturation {oxygen_saturation}% remains below 94%")
                criteria_used.append("Oxygen saturation 90-93%")
                explainability_vitals.append(f"Oxygen saturation at {oxygen_saturation}%")

        if invalid_ranges:
            risk_output = "Blocked - Invalid Vitals"
            pathway = "Final Triage Blocked"
            triggers = invalid_ranges
            criteria_used.append("Invalid vital-sign range blocks final triage")
        elif emergency_gate_active:
            risk_output = "High Risk"
            pathway = "Immediate ED Escalation"
            triggers = emergency_gates + [reason for reason in high_reasons if reason not in emergency_gates]
            criteria_used.append("Emergency gate active before scoring")
        elif critical_missing:
            risk_output = "Blocked - Critical Data Missing"
            pathway = "Final Triage Blocked"
            triggers = [f"Missing critical data: {item}" for item in critical_missing]
            criteria_used.append("Critical data missing blocks final triage")
        elif high_reasons:
            risk_output = "High Risk"
            pathway = "Immediate ED Escalation"
            triggers = high_reasons
        elif medium_reasons:
            risk_output = "Medium Risk"
            pathway = "ED High-Acuity Admission" if target == "Risk of urgent escalation during the same visit" else "Ward Admission"
            triggers = medium_reasons
        else:
            risk_output = "Low Risk"
            pathway = "Routine Evaluation"
            low_reasons.append("No major red flags, severe symptoms, or emergency vital sign abnormalities were detected from the entered data")
            if previous_admissions == 0:
                low_reasons.append("No previous admissions were recorded")
            triggers = low_reasons
            criteria_used.append("No High Risk criteria")
            criteria_used.append("No target-specific Medium Risk criteria")

        criteria_used = list(dict.fromkeys(criteria_used))

        # Presentation-aligned deterioration score, alert, and action
        score = 20
        score += min(45, len(high_reasons) * 18)
        if not high_reasons:
            score += min(28, len(medium_reasons) * 8)
        if previous_admissions is not None:
            score += min(10, previous_admissions * 2)
        score = max(0, min(100, score))

        if invalid_ranges:
            score = 0
            early_warning_alert = "SAFETY BLOCK: invalid vital-sign value entered"
            recommended_action = "Correct the invalid vital-sign entry before any final triage decision is saved"
        elif emergency_gate_active:
            score = max(score, 90)
            early_warning_alert = "EMERGENCY GATE ACTIVE: immediate ED escalation required before routine scoring"
            recommended_action = "Immediate ED physician review, resuscitation/critical-care evaluation, and no discharge pathway"
        elif critical_missing:
            score = 0
            early_warning_alert = "SAFETY BLOCK: missing critical triage data"
            recommended_action = "Complete critical data before final triage; do not use a low-risk or discharge pathway"
        elif high_reasons:
            score = max(score, 78)
            early_warning_alert = "Immediate ED escalation alert"
            recommended_action = "Immediate ED physician review, urgent intervention, and escalation planning"
        elif medium_reasons:
            score = max(score, 48)
            early_warning_alert = "Urgent review dashboard flag"
            recommended_action = "Prompt physician assessment, close monitoring, and escalation if the patient worsens"
        else:
            score = min(score, 35)
            early_warning_alert = "No immediate early warning alert"
            recommended_action = "Routine clinician evaluation with standard monitoring and follow-up planning"

        if target == "Risk of urgent escalation during the same visit" and risk_output == "High Risk":
            recommended_action = "Immediate ED escalation now: physician review, critical-care readiness, and same-visit escalation pathway"
        elif target == "Risk of 30-day hospitalization/readmission" and risk_output == "Medium Risk":
            recommended_action = "Clinician review with admission/readmission prevention planning and closer follow-up"
        elif target == "Risk of 30-day hospitalization/readmission" and risk_output == "High Risk":
            recommended_action = "Immediate ED escalation with likely admission and high-intensity follow-up planning"

        risk_score_text = "Blocked" if risk_output.startswith("Blocked") else f"{score}/100"

        assessed_fields = [age, systolic_bp, diastolic_bp, heart_rate, respiratory_rate, oxygen_saturation, temperature, previous_admissions]
        filled_count = sum(value is not None for value in assessed_fields)
        filled_count += 1 if medication_history_text else 0
        filled_count += 1 if comorbidities_text else 0
        filled_count += 1 if triage_notes else 0
        total_count = len(assessed_fields) + 3
        completeness_score = int((filled_count / total_count) * 100)

        if invalid_ranges or critical_missing:
            confidence_level = "Low Confidence"
        elif not critical_missing and completeness_score >= 85:
            confidence_level = "High Confidence"
        elif len(critical_missing) <= 2 and completeness_score >= 65:
            confidence_level = "Moderate Confidence"
        else:
            confidence_level = "Low Confidence"

        local_ed_acuity = determine_local_ed_acuity(
            emergency_gates=emergency_gates,
            risk_output=risk_output,
            missing_critical=critical_missing,
            invalid_ranges=invalid_ranges,
        )
        clinician_confirmed = self.final_decision_confirmed_var.get() if hasattr(self, "final_decision_confirmed_var") else False
        discharge_locked, discharge_lock_reasons = discharge_is_locked(
            {
                "emergency_gate_active": emergency_gate_active,
                "missing_critical": critical_missing,
                "invalid_ranges": invalid_ranges,
                "risk_output": risk_output,
            },
            clinician_confirmed=clinician_confirmed,
        )
        safety_lock_status = "LOCKED: " + "; ".join(discharge_lock_reasons) if discharge_locked else "Unlocked after complete data and clinician confirmation"

        evidence_lines = [
            f"Clinical setting: {ENTRY_POINT}",
            f"Target output: {target}",
            f"Prediction timing: {PREDICTION_TIMING_TEXT}",
            f"Primary users: {PREDICTION_USERS_TEXT}",
            f"Clinician oversight: {CLINICIAN_OVERSIGHT_TEXT}",
            f"Suggested result: {risk_output}",
            f"Risk score: {risk_score_text}",
            f"Early warning alert: {early_warning_alert}",
            f"Recommended action: {recommended_action}",
            f"Suggested pathway: {pathway}",
            f"Local ED acuity: {local_ed_acuity}",
            f"Emergency gate active: {'Yes' if emergency_gate_active else 'No'}",
            f"Do-not-discharge safety lock: {safety_lock_status}",
            f"Clinician confirmation: {'Confirmed' if clinician_confirmed else 'Pending'}",
            f"Confidence level: {confidence_level}",
            f"Triage completeness score: {completeness_score}%",
            "",
            "Criteria matched:",
        ]
        if criteria_used:
            evidence_lines.extend([f"- {item}" for item in criteria_used])
        else:
            evidence_lines.append("- No criteria matched from the entered data")

        evidence_lines.append("")
        evidence_lines.append("Why this result was suggested:")
        evidence_lines.extend([f"- {reason}" for reason in triggers])

        evidence_lines.append("")
        evidence_lines.append("Explainability panel:")
        evidence_lines.append("- Symptoms contributing to this output:")
        if explainability_symptoms:
            evidence_lines.extend([f"  • {item}" for item in explainability_symptoms])
        else:
            evidence_lines.append("  • No positive symptoms contributed")
        evidence_lines.append("- Vital sign findings contributing to this output:")
        if explainability_vitals:
            evidence_lines.extend([f"  • {item}" for item in explainability_vitals])
        else:
            evidence_lines.append("  • No major vital sign abnormalities contributed")
        evidence_lines.append("- History factors contributing to this output:")
        if explainability_history:
            evidence_lines.extend([f"  • {item}" for item in explainability_history])
        else:
            evidence_lines.append("  • No history factors contributed")
        evidence_lines.append("- Red flags contributing to this output:")
        if explainability_red_flags:
            evidence_lines.extend([f"  • {item}" for item in explainability_red_flags])
        else:
            evidence_lines.append("  • No red flags contributed")

        evidence_lines.append("")
        evidence_lines.append("Symptoms captured:")
        if symptoms_yes:
            evidence_lines.extend([f"- {item}" for item in symptoms_yes])
        else:
            evidence_lines.append("- No positive symptom flags were entered")
        evidence_lines.append("")
        evidence_lines.append("Red flags captured:")
        if red_flags_yes:
            evidence_lines.extend([f"- {item}" for item in red_flags_yes])
        else:
            evidence_lines.append("- No red flags were entered")
        evidence_lines.append("")
        evidence_lines.append("History factors captured:")
        if history_yes:
            evidence_lines.extend([f"- {item}" for item in history_yes])
        else:
            evidence_lines.append("- No additional history factors were entered")
        evidence_lines.append("")
        evidence_lines.append("Key captured values:")
        evidence_lines.append(f"- Entry point: {ENTRY_POINT}")
        evidence_lines.append(f"- BP: {self.case_vars['systolic_bp'].get()}/{self.case_vars['diastolic_bp'].get()}")
        evidence_lines.append(f"- HR: {self.case_vars['heart_rate'].get()}")
        evidence_lines.append(f"- RR: {self.case_vars['respiratory_rate'].get()}")
        evidence_lines.append(f"- SpO2: {self.case_vars['oxygen_saturation'].get()}")
        evidence_lines.append(f"- Temperature: {self.case_vars['temperature'].get()}")
        evidence_lines.append(f"- Previous admissions: {self.case_vars['previous_admissions'].get()}")
        evidence_lines.append("")
        evidence_lines.append("Emergency gates before scoring:")
        if emergency_gates:
            evidence_lines.extend([f"- {item}" for item in emergency_gates])
        else:
            evidence_lines.append("- No emergency gate active")
        evidence_lines.append("")
        evidence_lines.append("Invalid vital-sign ranges:")
        if invalid_ranges:
            evidence_lines.extend([f"- {item}" for item in invalid_ranges])
        else:
            evidence_lines.append("- No invalid vital-sign range detected")
        evidence_lines.append("")
        evidence_lines.append("Missing critical fields:")
        if critical_missing:
            evidence_lines.extend([f"- {item}" for item in critical_missing])
        else:
            evidence_lines.append("- No missing critical fields detected")
        if missing:
            evidence_lines.append("")
            evidence_lines.append("Other missing or invalid fields:")
            evidence_lines.extend([f"- {item}" for item in missing])

        red_flags_detected = list(red_flags_yes)
        if systolic_bp is not None and systolic_bp < 90:
            red_flags_detected.append(f"Hypotension: systolic BP {systolic_bp}")
        if oxygen_saturation is not None and oxygen_saturation < 90:
            red_flags_detected.append(f"Low oxygen saturation: {oxygen_saturation}%")
        if respiratory_rate is not None and respiratory_rate >= 30:
            red_flags_detected.append(f"High respiratory rate: {respiratory_rate}")
        if heart_rate is not None and heart_rate >= 130:
            red_flags_detected.append(f"Severe tachycardia: HR {heart_rate}")

        evidence = {
            "risk_output": risk_output,
            "final_pathway": pathway,
            "risk_score_text": risk_score_text,
            "early_warning_alert": early_warning_alert,
            "recommended_action": recommended_action,
            "triggers": triggers,
            "missing": missing,
            "missing_critical": critical_missing,
            "invalid_ranges": invalid_ranges,
            "emergency_gates": emergency_gates,
            "emergency_gate_active": emergency_gate_active,
            "local_ed_acuity": local_ed_acuity,
            "discharge_locked": discharge_locked,
            "discharge_lock_reasons": discharge_lock_reasons,
            "safety_lock_status": safety_lock_status,
            "clinician_confirmation_status": "Confirmed" if clinician_confirmed else "Pending",
            "criteria_used": criteria_used,
            "confidence_level": confidence_level,
            "completeness_score": completeness_score,
            "evidence_text": "\n".join(evidence_lines),
            "red_flags_detected": red_flags_detected,
        }
        return self.apply_alert_fatigue_control(evidence)

    def parse_previous_admissions(self, raw_value):
        """Accept numeric count or yes/no text and convert to an integer."""
        text_value = (raw_value or "").strip().lower()
        if text_value in ("", "no", "n", "none"):
            return 0
        if text_value in ("yes", "y"):
            return 1
        return int(text_value)

    def validate_case_inputs(self, for_update=False):
        values = {key: var.get().strip() for key, var in self.case_vars.items() if key != "triage_nurse_notes"}
        values["triage_nurse_notes"] = self.case_notes_text.get("1.0", "end").strip()
        required = [
            "patient_id", "patient_name", "age", "sex", "arrival_time",
            "systolic_bp", "diastolic_bp", "heart_rate", "respiratory_rate", "oxygen_saturation", "temperature"
        ]
        for field in required:
            if not values[field]:
                messagebox.showwarning("Validation", f"{field.replace('_', ' ').title()} is required.")
                return None

        integer_fields = ["age", "systolic_bp", "diastolic_bp", "heart_rate", "respiratory_rate", "oxygen_saturation"]
        for field in integer_fields:
            text_value = values.get(field, "0") or "0"
            try:
                values[field] = int(text_value)
            except ValueError:
                messagebox.showwarning("Validation", f"{field.replace('_', ' ').title()} must be a whole number.")
                return None

        try:
            values["previous_admissions"] = self.parse_previous_admissions(values.get("previous_admissions", "0"))
        except ValueError:
            messagebox.showwarning("Validation", "Previous admissions must be a number or yes/no.")
            return None

        try:
            values["temperature"] = float(values["temperature"])
        except ValueError:
            messagebox.showwarning("Validation", "Temperature must be numeric.")
            return None

        vital_range_errors = validate_vital_ranges_snapshot(values)
        if vital_range_errors:
            messagebox.showerror(
                "Vital Range Validation",
                "Final triage cannot be saved until these entries are corrected:\n\n- " + "\n- ".join(vital_range_errors),
            )
            return None

        for index, case in enumerate(self.patient_cases):
            if case["patient_id"].strip().lower() == values["patient_id"].lower():
                if not for_update or index != self.selected_case_index:
                    messagebox.showwarning("Validation", "Duplicate Patient ID is not allowed.")
                    return None

        values["target_output"] = self.get_active_prediction_target()
        values["entry_point"] = ENTRY_POINT
        values.update(self.get_predictive_use_case_metadata(values["target_output"]))
        evidence = self.build_risk_evidence()
        values["risk_output"] = self.override_risk_var.get().strip() or evidence["risk_output"]
        values["final_pathway"] = self.override_pathway_var.get().strip() or evidence["final_pathway"]
        values["evidence_text"] = evidence["evidence_text"]
        values["trigger_summary"] = "; ".join(evidence["triggers"])
        values["confidence_level"] = evidence["confidence_level"]
        values["completeness_score"] = evidence["completeness_score"]
        values["missing_critical_fields"] = "; ".join(evidence["missing_critical"])
        values["override_risk_output"] = self.override_risk_var.get().strip()
        values["override_final_pathway"] = self.override_pathway_var.get().strip()
        values["override_reason"] = self.override_reason_text.get("1.0", "end").strip()
        values["reviewer_role"] = self.user_role_var.get().strip() if hasattr(self, "user_role_var") else ""
        values["generated_document_type"] = self.document_type_var.get().strip() if hasattr(self, "document_type_var") else ""
        values["generated_document_text"] = self.documentation_text.get("1.0", "end").strip() if hasattr(self, "documentation_text") else ""
        values["doc_accuracy_checked"] = "Yes" if hasattr(self, "doc_accuracy_var") and self.doc_accuracy_var.get() else "No"
        values["doc_completeness_checked"] = "Yes" if hasattr(self, "doc_completeness_var") and self.doc_completeness_var.get() else "No"
        values["doc_traceability_checked"] = "Yes" if hasattr(self, "doc_traceability_var") and self.doc_traceability_var.get() else "No"
        values["doc_safety_checked"] = "Yes" if hasattr(self, "doc_safety_var") and self.doc_safety_var.get() else "No"
        values["alert_delivery_mode"] = self.alert_delivery_var.get().strip() if hasattr(self, "alert_delivery_var") else ""
        values["alert_fatigue_status"] = self.alert_fatigue_var.get().strip() if hasattr(self, "alert_fatigue_var") else ""
        values["local_ed_acuity"] = evidence.get("local_ed_acuity", "")
        values["safety_lock_status"] = evidence.get("safety_lock_status", "")
        values["clinician_confirmation_status"] = evidence.get("clinician_confirmation_status", "")
        values["emergency_gates"] = "; ".join(evidence.get("emergency_gates", []))
        values["invalid_ranges"] = "; ".join(evidence.get("invalid_ranges", []))

        if not (hasattr(self, "final_decision_confirmed_var") and self.final_decision_confirmed_var.get()):
            messagebox.showwarning(
                "Clinician Confirmation Required",
                "The final pathway is still pending clinician confirmation.\n\nClick 'Confirm Final Pathway' before saving this patient case.",
            )
            return None

        if evidence.get("invalid_ranges"):
            messagebox.showerror("Safety Block", "Invalid vital-sign ranges block final triage saving.")
            return None

        if evidence.get("missing_critical") and not evidence.get("emergency_gate_active"):
            messagebox.showwarning(
                "Safety Block",
                "Critical data are missing. Final triage saving is blocked until they are completed.\n\n- " + "\n- ".join(evidence.get("missing_critical", [])),
            )
            return None

        if "discharge" in values["final_pathway"].lower() and evidence.get("discharge_locked"):
            messagebox.showerror(
                "Do Not Discharge Lock",
                "The discharge pathway is locked because safety conditions are not cleared.\n\n" + evidence.get("safety_lock_status", ""),
            )
            return None

        return values

    def suggest_risk_output(self):
        if hasattr(self, "final_decision_confirmed_var"):
            self.final_decision_confirmed_var.set(False)
        if hasattr(self, "clinician_confirmation_var"):
            self.clinician_confirmation_var.set("Pending clinician confirmation")
        evidence = self.build_risk_evidence()
        self.case_vars["risk_output"].set(evidence["risk_output"])
        self.case_vars["final_pathway"].set(evidence["final_pathway"])
        self.update_decision_support_panels(evidence)
        if evidence["missing_critical"]:
            self.update_status("Provisional risk output suggested; review missing critical fields and confidence level")
        elif evidence["missing"]:
            self.update_status("Risk output suggested; review non-critical missing fields")
        else:
            self.update_status("Risk output suggested with explainability and confidence")
        alert_notes = evidence.get("alert_fatigue_status", "AI-assisted triage result displayed to clinician")
        clinician_action = "review pending"
        if evidence.get("suppression_active"):
            clinician_action = "duplicate suppressed"
        elif evidence.get("interruptive_alert_active"):
            clinician_action = "interruptive alert active"
        elif evidence.get("risk_output") == "Medium Risk":
            clinician_action = "dashboard flag active"
        self.append_audit_log(
            event_type="AI_RESULT_DISPLAYED",
            clinician_action=clinician_action,
            notes=alert_notes,
        )
               # Direct n8n emergency alert after risk classification
        try:
            emergency_gates = evidence.get("emergency_gates", [])
            if isinstance(emergency_gates, list):
                emergency_gates_text = "; ".join(emergency_gates)
            else:
                emergency_gates_text = str(emergency_gates or "")

            risk_output_text = str(evidence.get("risk_output", "")).strip()

            send_to_doctor = risk_output_text in ["Low Risk", "Medium Risk", "High Risk"]

            if send_to_doctor:
                emergency_payload = {
                    "alert_id": f"{self.case_vars['patient_id'].get().strip()}-{datetime.now().strftime('%Y%m%d%H%M%S')}",
                    "timestamp": datetime.now().isoformat(timespec="seconds"),
                    "event_type": "AI_RESULT_DISPLAYED",
                    "user_role": self.user_role_var.get().strip() if hasattr(self, "user_role_var") else "",

                    "patient_id": self.case_vars["patient_id"].get().strip(),

                    "risk_output": evidence.get("risk_output"),
                    "review_priority": (
                    "urgent" if risk_output_text == "High Risk" or emergency_gates_text.strip()
                    else "same_day_review" if risk_output_text == "Medium Risk"
                    else "routine_review"
                    ),
                    "doctor_review_required": "yes",
                    "risk_score": f"{evidence.get('risk_score', '')}/100",
                    "final_pathway": evidence.get("final_pathway"),
                    "recommended_action": evidence.get("recommended_action"),
                    "emergency_gates": emergency_gates_text,

                    "age": self.case_vars["age"].get().strip(),
                    "systolic_bp": self.case_vars["systolic_bp"].get().strip(),
                    "diastolic_bp": self.case_vars["diastolic_bp"].get().strip(),
                    "heart_rate": self.case_vars["heart_rate"].get().strip(),
                    "respiratory_rate": self.case_vars["respiratory_rate"].get().strip(),
                    "oxygen_saturation": self.case_vars["oxygen_saturation"].get().strip(),
                    "temperature": self.case_vars["temperature"].get().strip(),
                }
                               
                self.last_alert_id = emergency_payload.get("alert_id")
                sent = send_emergency_alert_to_n8n(emergency_payload)
                               

                if sent:
                    self.update_status("Doctor review request sent to n8n.")
                else:
                    self.update_status("Doctor review request failed.")

        except Exception as alert_error:
            print(f"n8n direct alert error: {alert_error}")
        self.generate_case_summary()
    def check_doctor_decision(self):
        """
        Fetch the doctor's final pathway decision from n8n using the latest alert_id.
        """
        alert_id = getattr(self, "last_alert_id", "")

        if not alert_id:
            messagebox.showwarning(
                "No Alert ID",
                "No emergency alert_id found yet.\n\nGenerate a High Risk alert first."
            )
            return

        try:
            response = requests.get(
                N8N_DOCTOR_DECISION_LOOKUP_URL,
                params={"alert_id": alert_id},
                timeout=10
            )
            response.raise_for_status()
            decision = response.json()

        except Exception as error:
            messagebox.showerror(
                "Doctor Decision Error",
                f"Could not check doctor decision from n8n.\n\nError:\n{error}"
            )
            return

        if decision.get("status") != "found":
            messagebox.showinfo(
                "No Doctor Decision Yet",
                f"No confirmed doctor decision found yet for alert_id:\n{alert_id}"
            )
            return

        final_pathway = decision.get("final_pathway", "")
        decision_status = decision.get("decision_status", "")
        decision_time = decision.get("decision_time", "")

        if not final_pathway:
            messagebox.showwarning(
                "Missing Decision",
                "Doctor decision was found, but final_pathway is empty."
            )
            return

        if hasattr(self, "case_vars") and "final_pathway" in self.case_vars:
            self.case_vars["final_pathway"].set(final_pathway)

        if hasattr(self, "clinician_confirmation_var"):
            self.clinician_confirmation_var.set("Confirmed by doctor via n8n")

        if hasattr(self, "final_decision_confirmed_var"):
            self.final_decision_confirmed_var.set(True)

        self.append_audit_log(
            event_type="FINAL_DECISION_CONFIRMED",
            clinician_action="confirmed via n8n doctor decision",
            notes=f"Doctor decision imported from n8n. Final pathway: {final_pathway}. Decision time: {decision_time}. Status: {decision_status}",
            document_type="doctor_decision"
        )

        if hasattr(self, "refresh_audit_tree"):
            self.refresh_audit_tree()

        self.update_status(f"Doctor decision imported: {final_pathway}")

        messagebox.showinfo(
            "Doctor Decision Imported",
            f"Doctor decision found and applied.\n\nFinal Pathway:\n{final_pathway}\n\nStatus:\n{decision_status}"
        )
      
    def generate_case_summary(self):
        patient_name = self.case_vars["patient_name"].get().strip() or "Unnamed patient"
        evidence = self.build_risk_evidence()
        top_reasons = "; ".join(evidence["triggers"][:3]) if evidence["triggers"] else "No major triggers captured"
        active_target = self.get_active_prediction_target()
        display_risk = self.override_risk_var.get().strip() or evidence["risk_output"]
        display_pathway = self.override_pathway_var.get().strip() or evidence["final_pathway"]
        summary = (
            f"Patient: {patient_name} | Entry Point: {ENTRY_POINT} | Target Output: {active_target}\n"
            f"Key symptoms: SOB={self.case_vars['shortness_of_breath'].get()}, Edema={self.case_vars['edema'].get()}, Fatigue={self.case_vars['fatigue'].get()}, Chest pain={self.case_vars['chest_pain'].get()}\n"
            f"Vitals: BP={self.case_vars['systolic_bp'].get()}/{self.case_vars['diastolic_bp'].get()}, HR={self.case_vars['heart_rate'].get()}, RR={self.case_vars['respiratory_rate'].get()}, SpO2={self.case_vars['oxygen_saturation'].get()}, Temp={self.case_vars['temperature'].get()}\n"
            f"Red flags: Severe dyspnea={self.case_vars['severe_dyspnea'].get()}, Confusion={self.case_vars['confusion'].get()}, Pulmonary edema signs={self.case_vars['pulmonary_edema_signs'].get()}, Arrhythmia symptoms={self.case_vars['arrhythmia_symptoms'].get()}\n"
            f"Suggested output: {display_risk} | Pathway: {display_pathway}\n"
            f"Main reasons: {top_reasons}"
        )
        self.case_vars["risk_output"].set(display_risk)
        self.case_vars["final_pathway"].set(display_pathway)
        self.case_summary_text.delete("1.0", "end")
        self.case_summary_text.insert("1.0", summary)
        self.case_evidence_text.delete("1.0", "end")
        self.case_evidence_text.insert("1.0", evidence["evidence_text"])
        self.update_decision_support_panels(evidence)
        self.update_status("Case summary generated with supporting reasons")
    def add_case(self):
        values = self.validate_case_inputs(for_update=False)
        if not values:
            return
        self.patient_cases.append(values)
        self.refresh_case_tree()
        self.append_audit_log(
            event_type="PATIENT_REGISTERED",
            clinician_action="registered",
            notes="Patient registered in triage system",
        )
        self.append_audit_log(
            event_type="FINAL_DECISION_CONFIRMED",
            clinician_action="confirmed",
            notes="Patient case added and final decision saved",
        )
        self.clear_case_form()
        self.update_status("Patient case added")
        messagebox.showinfo("Added", "Patient case added successfully.")

    def update_case(self):
        if self.selected_case_index is None:
            messagebox.showwarning("Update", "Select a patient case to update.")
            return
        values = self.validate_case_inputs(for_update=True)
        if not values:
            return
        self.patient_cases[self.selected_case_index] = values
        self.refresh_case_tree()
        self.append_audit_log(
            event_type="FINAL_DECISION_CONFIRMED",
            clinician_action="confirmed",
            notes="Patient case updated and final decision saved",
        )
        self.clear_case_form()
        self.update_status("Patient case updated")
        messagebox.showinfo("Updated", "Patient case updated successfully.")

    def delete_case(self):
        if self.selected_case_index is None:
            messagebox.showwarning("Delete", "Select a patient case to delete.")
            return
        if messagebox.askyesno("Confirm Delete", "Delete the selected patient case?"):
            del self.patient_cases[self.selected_case_index]
            self.refresh_case_tree()
            self.clear_case_form()
            self.update_status("Patient case deleted")

    def refresh_case_tree(self):
        for item in self.case_tree.get_children():
            self.case_tree.delete(item)

        search = self.case_search_var.get().strip().lower()
        for case in self.patient_cases:
            combined = f"{case['patient_id']} {case['patient_name']}".lower()
            if search and search not in combined:
                continue
            tags = ()
            if case["risk_output"] == "High Risk":
                tags = ("high_risk",)
            elif case["risk_output"] == "Medium Risk":
                tags = ("medium_risk",)
            elif case["risk_output"] == "Low Risk":
                tags = ("low_risk",)
            self.case_tree.insert("", "end", values=(
                case["patient_id"], case["patient_name"], case["age"], case["sex"], case["arrival_time"], case["risk_output"], case["final_pathway"], case.get("target_output", self.project_target_var.get())
            ), tags=tags)
        self.refresh_patient_selector()

    def refresh_patient_selector(self):
        options = []
        self._patient_id_to_index = {}
        for idx, case in enumerate(self.patient_cases):
            self._patient_id_to_index[str(case.get("patient_id", "")).strip()] = idx
            display_name = case.get("patient_name", "") or "Unnamed patient"
            risk = case.get("risk_output", "")
            options.append(f"{case.get('patient_id', '')} | {display_name} | {risk}")
        self.patient_selector_combo["values"] = options

        if self.selected_case_index is not None and 0 <= self.selected_case_index < len(self.patient_cases):
            current = self.patient_cases[self.selected_case_index]
            current_display = f"{current.get('patient_id', '')} | {current.get('patient_name', '') or 'Unnamed patient'} | {current.get('risk_output', '')}"
            self.patient_selector_var.set(current_display)
        elif not options:
            self.patient_selector_var.set("")

    def load_case_by_index(self, index, sync_tree=True):
        if self._patient_load_in_progress:
            return
        if not self.patient_cases:
            return
        if index < 0 or index >= len(self.patient_cases):
            return

        self._patient_load_in_progress = True
        try:
            case = self.patient_cases[index]
            self.populate_case_form_from_record(case, selected_index=index)

            if sync_tree:
                patient_id = str(case.get("patient_id", "")).strip()
                for item_id in self.case_tree.get_children():
                    values = self.case_tree.item(item_id, "values")
                    if values and str(values[0]).strip() == patient_id:
                        self.case_tree.selection_set(item_id)
                        self.case_tree.focus(item_id)
                        self.case_tree.see(item_id)
                        break
        finally:
            self._patient_load_in_progress = False

    def on_patient_selector_change(self, event=None):
        if self._patient_load_in_progress:
            return
        selected_text = self.patient_selector_var.get().strip()
        if not selected_text:
            return
        patient_id = selected_text.split("|", 1)[0].strip()
        index = self._patient_id_to_index.get(patient_id)
        if index is not None:
            self.load_case_by_index(index)

    def load_previous_patient(self):
        if not self.patient_cases:
            messagebox.showwarning("Navigation", "No saved patients available.")
            return
        if self.selected_case_index is None:
            self.load_case_by_index(0)
            return
        new_index = self.selected_case_index - 1
        if new_index < 0:
            new_index = len(self.patient_cases) - 1
        self.load_case_by_index(new_index)

    def load_next_patient(self):
        if not self.patient_cases:
            messagebox.showwarning("Navigation", "No saved patients available.")
            return
        if self.selected_case_index is None:
            self.load_case_by_index(0)
            return
        new_index = self.selected_case_index + 1
        if new_index >= len(self.patient_cases):
            new_index = 0
        self.load_case_by_index(new_index)

    def populate_case_form_from_record(self, case_record, selected_index=None):
        self.selected_case_index = selected_index
        for key, var in self.case_vars.items():
            if key == "triage_nurse_notes":
                continue
            var.set(case_record.get(key, ""))
        self.case_notes_text.delete("1.0", "end")
        self.case_notes_text.insert("1.0", case_record.get("triage_nurse_notes", ""))
        self.override_risk_var.set(case_record.get("override_risk_output", ""))
        self.override_pathway_var.set(case_record.get("override_final_pathway", ""))
        self.override_reason_text.delete("1.0", "end")
        self.override_reason_text.insert("1.0", case_record.get("override_reason", ""))
        if hasattr(self, "case_target_var"):
            self.case_target_var.set(case_record.get("target_output", self.project_target_var.get()))
        if hasattr(self, "user_role_var"):
            self.user_role_var.set(case_record.get("reviewer_role", "Emergency Physician") or "Emergency Physician")
        if hasattr(self, "document_type_var"):
            self.document_type_var.set(case_record.get("generated_document_type", DOCUMENT_TYPES[0]) or DOCUMENT_TYPES[0])
        if hasattr(self, "documentation_text"):
            self.documentation_text.delete("1.0", "end")
            self.documentation_text.insert("1.0", case_record.get("generated_document_text", ""))
        if hasattr(self, "doc_accuracy_var"):
            self.doc_accuracy_var.set(case_record.get("doc_accuracy_checked", "No") == "Yes")
            self.doc_completeness_var.set(case_record.get("doc_completeness_checked", "No") == "Yes")
            self.doc_traceability_var.set(case_record.get("doc_traceability_checked", "No") == "Yes")
            self.doc_safety_var.set(case_record.get("doc_safety_checked", "No") == "Yes")
        if hasattr(self, "alert_delivery_var"):
            self.alert_delivery_var.set(case_record.get("alert_delivery_mode", "Not calculated") or "Not calculated")
        if hasattr(self, "alert_fatigue_var"):
            self.alert_fatigue_var.set(case_record.get("alert_fatigue_status", "No alert has been generated yet") or "No alert has been generated yet")
        display_name = case_record.get("patient_name", "") or "Unnamed patient"
        self.patient_selector_var.set(f"{case_record.get('patient_id', '')} | {display_name} | {case_record.get('risk_output', '')}")
        self.generate_case_summary()

    def on_case_select(self, event=None):
        if self._patient_load_in_progress:
            return
        selection = self.case_tree.selection()
        if not selection:
            return
        item_id = selection[0]
        values = self.case_tree.item(item_id, "values")
        patient_id = str(values[0]).strip()
        index = self._patient_id_to_index.get(patient_id)
        if index is not None:
            self.load_case_by_index(index, sync_tree=False)
    def clear_case_form(self):
        defaults = {
            "arrival_time": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "risk_output": "Medium Risk",
            "final_pathway": "Final Triage Blocked",
        }
        yes_no_fields = {"shortness_of_breath", "edema", "fatigue", "chest_pain", "severe_dyspnea", "confusion", "pulmonary_edema_signs", "arrhythmia_symptoms"}
        for key, var in self.case_vars.items():
            if key in defaults:
                var.set(defaults[key])
            elif key == "sex":
                var.set("")
            elif key in yes_no_fields:
                var.set("No")
            else:
                var.set("")
        self.case_notes_text.delete("1.0", "end")
        self.case_summary_text.delete("1.0", "end")
        self.case_evidence_text.delete("1.0", "end")
        if hasattr(self, "missing_critical_text"):
            self.missing_critical_text.delete("1.0", "end")
        if hasattr(self, "red_flags_text"):
            self.red_flags_text.delete("1.0", "end")
        if hasattr(self, "override_risk_var"):
            self.override_risk_var.set("")
        if hasattr(self, "override_pathway_var"):
            self.override_pathway_var.set("")
        if hasattr(self, "override_reason_text"):
            self.override_reason_text.delete("1.0", "end")
        if hasattr(self, "output_risk_var"):
            self.output_risk_var.set("Not calculated")
            self.output_pathway_var.set("Not calculated")
            self.set_risk_style("")
        if hasattr(self, "human_loop_var"):
            self.human_loop_var.set("Human-in-the-Loop: Final decision requires clinician review")
        if hasattr(self, "alert_delivery_var"):
            self.alert_delivery_var.set("Not calculated")
        if hasattr(self, "alert_fatigue_var"):
            self.alert_fatigue_var.set("No alert has been generated yet")
        if hasattr(self, "local_acuity_var"):
            self.local_acuity_var.set("Not calculated")
        if hasattr(self, "safety_lock_var"):
            self.safety_lock_var.set("Safety status not calculated")
        if hasattr(self, "triage_gate_var"):
            self.triage_gate_var.set("No emergency gate evaluated yet")
        if hasattr(self, "clinician_confirmation_var"):
            self.clinician_confirmation_var.set("Pending clinician confirmation")
        if hasattr(self, "final_decision_confirmed_var"):
            self.final_decision_confirmed_var.set(False)
        if hasattr(self, "acknowledge_alert_button"):
            self.acknowledge_alert_button.configure(state="disabled")
        if hasattr(self, "reset_alert_button"):
            self.reset_alert_button.configure(state="disabled")
        if hasattr(self, "confidence_var"):
            self.confidence_var.set("Not calculated")
            self.set_confidence_style("Not calculated")
        if hasattr(self, "case_target_var"):
            self.case_target_var.set(self.project_target_var.get())
        if hasattr(self, "user_role_var"):
            self.user_role_var.set("Emergency Physician")
        if hasattr(self, "document_type_var"):
            self.document_type_var.set(DOCUMENT_TYPES[0])
        if hasattr(self, "documentation_text"):
            self.documentation_text.delete("1.0", "end")
        if hasattr(self, "doc_accuracy_var"):
            self.doc_accuracy_var.set(False)
            self.doc_completeness_var.set(False)
            self.doc_traceability_var.set(False)
            self.doc_safety_var.set(False)
        self.selected_case_index = None
        self.alert_state_by_patient = {}
        self.current_alert_context = {}
        if hasattr(self, "patient_selector_var"):
            self.patient_selector_var.set("")
        self.update_patient_dashboard_cards({"missing_critical": []})

    def export_case_csv(self):
        if not self.patient_cases:
            messagebox.showwarning("Export", "No patient cases to export.")
            return
        filename = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")], title="Export Patient Cases CSV")
        if not filename:
            return
        fieldnames = [
            "patient_id", "patient_name", "age", "sex", "arrival_time",
            "shortness_of_breath", "edema", "fatigue", "chest_pain",
            "systolic_bp", "diastolic_bp", "heart_rate", "respiratory_rate", "oxygen_saturation", "temperature",
            "previous_admissions", "medication_history", "comorbidities",
            "severe_dyspnea", "confusion", "pulmonary_edema_signs", "arrhythmia_symptoms",
            "triage_nurse_notes", "target_output", "prediction_timing", "target_users", "output_type", "clinician_oversight", "entry_point", "trigger_summary", "confidence_level", "completeness_score", "missing_critical_fields", "evidence_text",
            "reviewer_role", "generated_document_type", "generated_document_text", "doc_accuracy_checked", "doc_completeness_checked", "doc_traceability_checked", "doc_safety_checked"
        ]
        try:
            with open(filename, "w", newline="", encoding="utf-8") as file:
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(self.patient_cases)
            messagebox.showinfo("Exported", "Patient cases CSV exported successfully.")
            self.update_status("Patient cases CSV exported")
        except Exception as error:
            messagebox.showerror("Export Error", f"Could not export patient cases CSV.\n\n{error}")

    def normalize_import_key(self, key):
        key = str(key).strip().lower()
        replacements = {
            "patient id": "patient_id",
            "patient name": "patient_name",
            "arrival time": "arrival_time",
            "shortness of breath": "shortness_of_breath",
            "chest pain": "chest_pain",
            "systolic bp": "systolic_bp",
            "diastolic bp": "diastolic_bp",
            "heart rate": "heart_rate",
            "respiratory rate": "respiratory_rate",
            "oxygen saturation": "oxygen_saturation",
            "previous admission": "previous_admissions",
            "previous admissions": "previous_admissions",
            "medication history": "medication_history",
            "pulmonary edema sign": "pulmonary_edema_signs",
            "pulmonary edema signs": "pulmonary_edema_signs",
            "arrhythmia symptom": "arrhythmia_symptoms",
            "arrhythmia symptoms": "arrhythmia_symptoms",
            "triage nurse notes": "triage_nurse_notes",
            "risk output": "risk_output",
            "final pathway": "final_pathway",
            "target output": "target_output",
            "prediction target": "target_output",
            "prediction timing": "prediction_timing",
            "target users": "target_users",
            "output type": "output_type",
            "clinician oversight": "clinician_oversight",
        }
        key = key.replace("-", " ").replace("/", " ")
        key = " ".join(key.split())
        key = replacements.get(key, key.replace(" ", "_"))
        return key

    def import_case_excel(self):
        filename = filedialog.askopenfilename(
            title="Import Patient Cases from Excel or CSV",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")],
        )
        if not filename:
            return

        rows = []
        try:
            if filename.lower().endswith(".csv"):
                with open(filename, "r", encoding="utf-8-sig", newline="") as file:
                    reader = csv.DictReader(file)
                    rows = list(reader)
            else:
                if not HAS_OPENPYXL:
                    messagebox.showerror("Import Error", "Excel import needs openpyxl. Install it with: pip install openpyxl")
                    return
                workbook = load_workbook(filename=filename)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                        rows.append({headers[i]: row[i] for i in range(len(headers))})
        except Exception as error:
            messagebox.showerror("Import Error", f"Could not import the selected file.\n\n{error}")
            return

        added_count = 0
        updated_count = 0
        skipped_count = 0
        affected_patient_ids = []

        for raw_row in rows:
            mapped = {}
            for key, value in raw_row.items():
                if key is None:
                    continue
                mapped[self.normalize_import_key(key)] = "" if value is None else str(value).strip()

            record = {key: "" for key in self.case_vars.keys()}
            for key in record.keys():
                if key in mapped:
                    record[key] = mapped[key]

            patient_id = record.get("patient_id", "").strip()
            if not patient_id:
                skipped_count += 1
                continue

            record.setdefault("sex", mapped.get("sex", ""))
            record.setdefault("arrival_time", mapped.get("arrival_time", datetime.now().strftime("%Y-%m-%d %H:%M")))
            record.setdefault("entry_point", ENTRY_POINT)
            record["risk_output"] = ""
            record["final_pathway"] = ""
            record["target_output"] = mapped.get("target_output", self.project_target_var.get())
            record.update(self.get_predictive_use_case_metadata(record["target_output"]))

            for key in self.case_vars:
                if key == "triage_nurse_notes":
                    continue
                self.case_vars[key].set(record.get(key, ""))
            self.case_notes_text.delete("1.0", "end")
            self.case_notes_text.insert("1.0", record.get("triage_nurse_notes", ""))
            if hasattr(self, "case_target_var"):
                self.case_target_var.set(record.get("target_output", self.project_target_var.get()))

            evidence = self.build_risk_evidence()
            record["risk_output"] = evidence["risk_output"]
            record["final_pathway"] = evidence["final_pathway"]
            record["evidence_text"] = evidence["evidence_text"]
            record["trigger_summary"] = "; ".join(evidence["triggers"])
            record["confidence_level"] = evidence["confidence_level"]
            record["completeness_score"] = evidence["completeness_score"]
            record["missing_critical_fields"] = "; ".join(evidence["missing_critical"])

            existing_index = next(
                (idx for idx, case in enumerate(self.patient_cases) if case["patient_id"].strip().lower() == patient_id.lower()),
                None,
            )
            if existing_index is None:
                self.patient_cases.append(record)
                added_count += 1
            else:
                self.patient_cases[existing_index] = record
                updated_count += 1

            affected_patient_ids.append(patient_id)

        self.refresh_case_tree()
        if affected_patient_ids:
            first_id = affected_patient_ids[0]
            for idx, case in enumerate(self.patient_cases):
                if case["patient_id"] == first_id:
                    self.populate_case_form_from_record(case, selected_index=idx)
                    break
        else:
            self.clear_case_form()

        self.update_status(f"Import complete: {added_count} added, {updated_count} updated, {skipped_count} skipped")
        messagebox.showinfo(
            "Import Complete",
            f"Added: {added_count}\nUpdated: {updated_count}\nSkipped: {skipped_count}\n\nDuplicate patient IDs are now updated instead of skipped."
        )

    def export_case_excel(self):
        if not self.patient_cases:
            messagebox.showwarning("Export", "No patient cases to export.")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("Export Error", "Excel export needs openpyxl. Install it with: pip install openpyxl")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Export Patient Cases Excel",
        )
        if not filename:
            return

        fieldnames = [
            "patient_id", "patient_name", "age", "sex", "arrival_time",
            "shortness_of_breath", "edema", "fatigue", "chest_pain",
            "systolic_bp", "diastolic_bp", "heart_rate", "respiratory_rate", "oxygen_saturation", "temperature",
            "previous_admissions", "medication_history", "comorbidities",
            "severe_dyspnea", "confusion", "pulmonary_edema_signs", "arrhythmia_symptoms",
            "triage_nurse_notes", "target_output", "prediction_timing", "target_users", "output_type", "clinician_oversight", "entry_point", "trigger_summary", "confidence_level", "completeness_score", "missing_critical_fields", "evidence_text",
            "reviewer_role", "generated_document_type", "generated_document_text", "doc_accuracy_checked", "doc_completeness_checked", "doc_traceability_checked", "doc_safety_checked"
        ]

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Patient Cases"
            for col_index, field in enumerate(fieldnames, start=1):
                sheet.cell(row=1, column=col_index, value=field)
            for row_index, case in enumerate(self.patient_cases, start=2):
                for col_index, field in enumerate(fieldnames, start=1):
                    sheet.cell(row=row_index, column=col_index, value=case.get(field, ""))
            workbook.save(filename)
            messagebox.showinfo("Exported", "Patient cases Excel exported successfully.")
            self.update_status("Patient cases Excel exported")
        except Exception as error:
            messagebox.showerror("Export Error", f"Could not export patient cases Excel.\n\n{error}")



if __name__ == "__main__":
    app = HFTriageApp()
    app.mainloop()
